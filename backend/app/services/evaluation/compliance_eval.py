from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import FileRecord, Meeting, Output
from app.services.domain.compliance.constants import CATEGORY_KEYWORDS


@dataclass(frozen=True)
class ComplianceEvalCase:
    case_id: str
    case_name: str
    source_folder: str
    expected_file_count: int
    expected_categories: dict[str, int]
    expected_facts: dict[str, Any]
    expected_template: dict[str, int]
    required_outputs: list[str]
    forbidden_loose_outputs: list[str]


def _case_file() -> Path:
    return Path(__file__).with_name("compliance_eval_cases.json")


def load_compliance_eval_cases(path: Path | None = None) -> list[ComplianceEvalCase]:
    raw = json.loads((path or _case_file()).read_text(encoding="utf-8"))
    cases: list[ComplianceEvalCase] = []
    for item in raw:
        cases.append(
            ComplianceEvalCase(
                case_id=str(item["case_id"]),
                case_name=str(item.get("case_name") or item["case_id"]),
                source_folder=str(item.get("source_folder") or ""),
                expected_file_count=int(item.get("expected_file_count") or 0),
                expected_categories=dict(item.get("expected_categories") or {}),
                expected_facts=dict(item.get("expected_facts") or {}),
                expected_template=dict(item.get("expected_template") or {}),
                required_outputs=list(item.get("required_outputs") or []),
                forbidden_loose_outputs=list(item.get("forbidden_loose_outputs") or []),
            )
        )
    return cases


def find_compliance_eval_case(
    meeting_code: str | None,
    cases: list[ComplianceEvalCase] | None = None,
) -> ComplianceEvalCase | None:
    code = str(meeting_code or "").strip().upper()
    if not code:
        return None
    for case in cases or load_compliance_eval_cases():
        if case.case_id.upper() == code:
            return case
        if code in case.source_folder.upper():
            return case
    return None


def _normalize(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _check(
    checks: list[dict[str, Any]],
    *,
    check_id: str,
    passed: bool,
    expected: Any,
    actual: Any,
    severity: str = "critical",
    message: str = "",
) -> None:
    checks.append(
        {
            "check_id": check_id,
            "passed": bool(passed),
            "severity": severity,
            "expected": expected,
            "actual": actual,
            "message": message,
        }
    )


FACT_RELATED_CATEGORIES: dict[str, list[str]] = {
    "actual_sign_in_count": ["sign_in_record"],
    "attendance_delta": ["sign_in_record", "observation_confirmation", "a1_meeting_export"],
    "attendance_source": ["sign_in_record", "meeting_screenshot"],
    "watch_record_count": ["sign_in_record", "meeting_screenshot"],
    "zoom_peak_count": ["meeting_screenshot"],
    "total_attendance_expression": ["observation_confirmation", "meeting_screenshot", "sign_in_record"],
    "planned_attendees": ["observation_confirmation", "a1_meeting_export"],
    "max_attendee_count": ["meeting_screenshot", "observation_confirmation"],
    "material_code": ["presentation_material"],
    "presentation_topic": ["presentation_material"],
}


def _file_item(raw: dict[str, Any]) -> dict[str, Any]:
    return {
        "file_id": raw.get("file_id") or raw.get("id"),
        "file_name": raw.get("file_name"),
        "file_type": raw.get("file_type"),
        "current_category": raw.get("document_category") or raw.get("current_category"),
        "confidence": raw.get("confidence"),
        "parse_status": raw.get("parse_status"),
    }


def _category_keyword_match(file_name: str, category: str) -> bool:
    lower = file_name.lower()
    return any(str(keyword).lower() in lower for keyword in CATEGORY_KEYWORDS.get(category, []))


def _candidate_files_for_category(files: list[dict[str, Any]], category: str) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for raw in files:
        file_name = str(raw.get("file_name") or "")
        current_category = raw.get("document_category") or raw.get("current_category")
        if current_category == category:
            item = _file_item(raw)
            item["reason"] = "current_category"
            candidates.append(item)
        elif _category_keyword_match(file_name, category):
            item = _file_item(raw)
            item["reason"] = "filename_keyword"
            candidates.append(item)
    return candidates[:8]


def _related_files_for_fact(files: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    related: list[dict[str, Any]] = []
    seen: set[str] = set()
    for category in FACT_RELATED_CATEGORIES.get(field, []):
        for item in _candidate_files_for_category(files, category):
            key = str(item.get("file_id") or item.get("file_name"))
            if key in seen:
                continue
            seen.add(key)
            item = dict(item)
            item["related_category"] = category
            related.append(item)
    return related[:8]


def _outputs_detail(actual: dict[str, Any]) -> list[dict[str, Any]]:
    return list(actual.get("outputs_detail") or [])


def _diagnose_failed_check(check: dict[str, Any], actual: dict[str, Any]) -> dict[str, Any]:
    check_id = str(check.get("check_id") or "")
    files = list(actual.get("files") or [])
    categories = dict(actual.get("file_categories") or {})

    if check_id == "file_count":
        return {
            "stage": "ingestion",
            "root_cause": "资料入库总数与基准不一致，需核对本场观察资料文件夹是否完整导入。",
            "category_counts": categories,
            "remediation": "回到资料页核对缺失或重复文件后重新运行分析。",
        }

    if check_id.startswith("category:"):
        category = check_id.split(":", 1)[1]
        candidates = _candidate_files_for_category(files, category)
        root_cause = (
            "存在文件名疑似属于该资料类型，但当前分类未计入目标类别，优先复核分类/OCR。"
            if candidates
            else "未找到可疑候选文件，可能是资料缺失或文件命名无法提示资料类型。"
        )
        return {
            "stage": "classification",
            "target_category": category,
            "expected_count": check.get("expected"),
            "actual_count": check.get("actual"),
            "root_cause": root_cause,
            "candidate_files": candidates,
            "remediation": "重新运行视觉识别与分类；若候选文件确认正确，应修正规则或资料类型。",
        }

    if check_id.startswith("fact:"):
        field = check_id.split(":", 1)[1]
        related_files = _related_files_for_fact(files, field)
        return {
            "stage": "fact_extraction",
            "field": field,
            "source_path": f"meeting.state_json.meeting_case.{field}",
            "expected_value": check.get("expected"),
            "actual_value": check.get("actual"),
            "root_cause": "事实字段与基准不一致，通常来自资料识别、OCR抽取或字段汇总规则错误。",
            "related_categories": FACT_RELATED_CATEGORIES.get(field, []),
            "related_files": related_files,
            "remediation": "打开关联文件核对原始内容，再重跑视觉识别/事实抽取或修正规则记忆。",
        }

    if check_id.startswith("template:"):
        field = check_id.split(":", 1)[1]
        template = dict(actual.get("template") or {})
        fixed_output = next(
            (item for item in _outputs_detail(actual) if item.get("output_type") == "fixed_template_excel"),
            {},
        )
        return {
            "stage": "template_generation",
            "field": field,
            "output_type": "fixed_template_excel",
            "file_name": template.get("file_name") or fixed_output.get("file_name"),
            "template_stats": template,
            "root_cause": "固定模板结构或字段填充数量不符合基准。",
            "remediation": "检查固定模板输出是否使用143列模板，并核对字段填充引擎的证据映射。",
        }

    if check_id.startswith("output:"):
        output_type = check_id.split(":", 1)[1]
        return {
            "stage": "delivery_scope",
            "missing_output_type": output_type,
            "available_outputs": [item.get("output_type") for item in _outputs_detail(actual)] or actual.get("outputs", []),
            "root_cause": "正式交付入口缺失。",
            "remediation": "重新生成交付物，确认固定模板Excel与ZIP归档包均已入库。",
        }

    if check_id == "output_scope:no_loose_outputs":
        loose_types = set(check.get("actual") or [])
        loose_outputs = [
            item for item in _outputs_detail(actual) if item.get("output_type") in loose_types
        ] or [{"output_type": item} for item in sorted(loose_types)]
        return {
            "stage": "delivery_scope",
            "loose_outputs": loose_outputs,
            "root_cause": "支撑文件作为散落交付入口出现，交付范围没有收敛到固定模板Excel与ZIP。",
            "remediation": "将Finding、证据索引、缺件清单等支撑文件纳入ZIP，仅保留主Excel和ZIP入口。",
        }

    return {
        "stage": "unknown",
        "root_cause": "该检查未通过，但暂无专用诊断规则。",
        "remediation": "查看完整评估快照与运行日志。",
    }


def evaluate_compliance_snapshot(
    case: ComplianceEvalCase,
    actual: dict[str, Any],
) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    categories = dict(actual.get("file_categories") or {})
    facts = dict(actual.get("facts") or {})
    template = dict(actual.get("template") or {})
    outputs = set(actual.get("outputs") or [])

    _check(
        checks,
        check_id="file_count",
        passed=int(actual.get("file_count") or 0) == case.expected_file_count,
        expected=case.expected_file_count,
        actual=actual.get("file_count"),
        message="资料总数必须与真实样本基准一致",
    )

    for category, expected_count in case.expected_categories.items():
        _check(
            checks,
            check_id=f"category:{category}",
            passed=int(categories.get(category) or 0) == int(expected_count),
            expected=expected_count,
            actual=categories.get(category, 0),
            message=f"{category} 分类数量必须稳定",
        )

    for key, expected_value in case.expected_facts.items():
        actual_value = facts.get(key)
        _check(
            checks,
            check_id=f"fact:{key}",
            passed=_normalize(actual_value) == _normalize(expected_value),
            expected=expected_value,
            actual=actual_value,
            message=f"{key} 必须来自资料事实链，不允许靠默认值或错误兜底",
        )

    for key, expected_value in case.expected_template.items():
        actual_value = template.get(key)
        _check(
            checks,
            check_id=f"template:{key}",
            passed=int(actual_value or 0) == int(expected_value),
            expected=expected_value,
            actual=actual_value,
            message=f"固定模板 {key} 必须满足基准",
        )

    for output_type in case.required_outputs:
        _check(
            checks,
            check_id=f"output:{output_type}",
            passed=output_type in outputs,
            expected=True,
            actual=output_type in outputs,
            message="正式交付入口必须存在",
        )

    loose = sorted(output for output in case.forbidden_loose_outputs if output in outputs)
    _check(
        checks,
        check_id="output_scope:no_loose_outputs",
        passed=not loose,
        expected=[],
        actual=loose,
        severity="warning",
        message="Finding、证据索引等支撑文件应在 ZIP 内，不应作为正式散落交付入口",
    )

    for item in checks:
        if not item["passed"]:
            item["diagnosis"] = _diagnose_failed_check(item, actual)

    critical_failures = sum(
        1 for item in checks if not item["passed"] and item.get("severity") == "critical"
    )
    warning_failures = sum(
        1 for item in checks if not item["passed"] and item.get("severity") == "warning"
    )
    return {
        "case_id": case.case_id,
        "case_name": case.case_name,
        "passed": critical_failures == 0,
        "critical_failures": critical_failures,
        "warning_failures": warning_failures,
        "checks": checks,
    }


def compact_compliance_evaluation(
    report: dict[str, Any],
    *,
    max_failed_checks: int = 8,
) -> dict[str, Any]:
    failed_checks = [
        {
            "check_id": item.get("check_id"),
            "severity": item.get("severity"),
            "expected": item.get("expected"),
            "actual": item.get("actual"),
            "message": item.get("message"),
            "diagnosis": item.get("diagnosis"),
        }
        for item in report.get("checks", [])
        if not item.get("passed")
    ][:max_failed_checks]
    compact = {
        "status": report.get("status", "completed"),
        "case_id": report.get("case_id"),
        "case_name": report.get("case_name"),
        "meeting_code": report.get("meeting_code"),
        "passed": report.get("passed"),
        "critical_failures": report.get("critical_failures", 0),
        "warning_failures": report.get("warning_failures", 0),
        "total_checks": report.get("total_checks", len(report.get("checks", []))),
        "passed_checks": report.get("passed_checks"),
        "failed_checks": failed_checks,
        "generated_at": report.get("generated_at"),
    }
    if report.get("reason"):
        compact["reason"] = report["reason"]
    return compact


def _fixed_template_stats(path: Path) -> dict[str, int]:
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        header_row = 0
        columns = 0
        for row_idx in range(1, min(ws.max_row, 5) + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            count = sum(1 for value in values if value not in (None, ""))
            if count > columns:
                header_row = row_idx
                columns = count
        value_row = header_row + 1 if header_row else 0
        filled = 0
        if value_row:
            filled = sum(
                1
                for col in range(1, ws.max_column + 1)
                if ws.cell(value_row, col).value not in (None, "")
            )
        return {"columns": columns, "filled_row2": filled}
    finally:
        wb.close()


def build_db_compliance_snapshot(
    db: Session,
    project_id: str,
    meeting_id: str,
) -> dict[str, Any]:
    meeting = db.query(Meeting).filter_by(project_id=project_id, id=meeting_id).first()
    if not meeting:
        raise ValueError("meeting not found")

    files = (
        db.query(FileRecord)
        .filter_by(project_id=project_id, meeting_id=meeting_id)
        .order_by(FileRecord.file_name.asc())
        .all()
    )
    outputs = db.query(Output).filter_by(project_id=project_id, meeting_id=meeting_id).all()
    categories: dict[str, int] = {}
    for file_record in files:
        categories[file_record.document_category] = categories.get(file_record.document_category, 0) + 1

    template: dict[str, int] = {}
    for output in outputs:
        if output.output_type == "fixed_template_excel":
            template = _fixed_template_stats(Path(output.storage_path))
            break

    output_order = {"fixed_template_excel": 0, "deliverable_package": 1}
    outputs_detail = sorted(
        [
            {
                "output_id": output.id,
                "output_type": output.output_type,
                "file_name": output.file_name,
                "path_exists": Path(output.storage_path).exists(),
            }
            for output in outputs
        ],
        key=lambda item: (output_order.get(str(item["output_type"]), 99), str(item["output_type"])),
    )

    return {
        "file_count": len(files),
        "file_categories": categories,
        "files": [
            {
                "file_id": file_record.id,
                "file_name": file_record.file_name,
                "file_type": file_record.file_type,
                "document_category": file_record.document_category,
                "confidence": file_record.confidence,
                "parse_status": file_record.parse_status,
            }
            for file_record in files
        ],
        "facts": dict((meeting.state_json or {}).get("meeting_case") or {}),
        "template": template,
        "outputs": sorted({output.output_type for output in outputs}),
        "outputs_detail": outputs_detail,
    }


def run_db_compliance_evaluation(
    db: Session,
    project_id: str,
    meeting_id: str,
) -> dict[str, Any]:
    meeting = db.query(Meeting).filter_by(project_id=project_id, id=meeting_id).first()
    if not meeting:
        raise ValueError("meeting not found")

    state = dict(meeting.state_json or {})
    meeting_case = dict(state.get("meeting_case") or {})
    meeting_code = str(meeting_case.get("meeting_code") or meeting.meeting_code or "")
    generated_at = datetime.now(timezone.utc).isoformat()
    case = find_compliance_eval_case(meeting_code)
    if not case:
        return {
            "status": "skipped",
            "reason": "no_baseline_case",
            "meeting_code": meeting_code,
            "passed": None,
            "critical_failures": 0,
            "warning_failures": 0,
            "checks": [],
            "total_checks": 0,
            "passed_checks": 0,
            "generated_at": generated_at,
        }

    snapshot = build_db_compliance_snapshot(db, project_id, meeting_id)
    report = evaluate_compliance_snapshot(case, snapshot)
    report.update(
        {
            "status": "completed",
            "meeting_code": meeting_code,
            "snapshot": snapshot,
            "total_checks": len(report["checks"]),
            "passed_checks": sum(1 for item in report["checks"] if item.get("passed")),
            "generated_at": generated_at,
        }
    )
    return report

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.comments
import openpyxl.styles
import openpyxl.utils

from app.services.classifier import detect_header_row, normalize_field


def parse_excel(file_path: Path) -> dict[str, Any]:
    if file_path.suffix.lower() == ".csv":
        return _parse_csv(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_out: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))
        if not raw_rows:
            continue

        header_idx = detect_header_row(raw_rows)
        headers = [
            str(c).strip() if c is not None else f"col_{i}"
            for i, c in enumerate(raw_rows[header_idx])
        ]
        columns = []
        for i, h in enumerate(headers):
            std, conf = normalize_field(h)
            columns.append(
                {
                    "name": h,
                    "standard_field": std,
                    "type": _infer_type(std, h),
                    "column_letter": openpyxl.utils.get_column_letter(i + 1),
                    "normalize_confidence": conf,
                }
            )

        rows = []
        for r_idx, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            values = {}
            for col_idx, col in enumerate(columns):
                val = row[col_idx] if col_idx < len(row) else None
                values[col["name"]] = _normalize_value(val, col["type"])
            rows.append({"row_number": r_idx, "values": values})

        sheets_out.append(
            {
                "sheet_name": sheet_name,
                "header_row": header_idx + 1,
                "pre_header_rows": raw_rows[:header_idx],
                "columns": columns,
                "rows": rows,
            }
        )

    return {"file_type": "excel", "sheets": sheets_out}


def _parse_csv(file_path: Path) -> dict[str, Any]:
    import csv

    with file_path.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        raw_rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not raw_rows:
        return {"file_type": "excel", "sheets": []}

    header_idx = detect_header_row(raw_rows)
    headers = [c.strip() or f"col_{i}" for i, c in enumerate(raw_rows[header_idx])]
    columns = []
    for i, h in enumerate(headers):
        std, conf = normalize_field(h)
        columns.append(
            {
                "name": h,
                "standard_field": std,
                "type": _infer_type(std, h),
                "column_letter": openpyxl.utils.get_column_letter(i + 1),
                "normalize_confidence": conf,
            }
        )
    rows = []
    for r_idx, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
        values = {}
        for col_idx, col in enumerate(columns):
            val = row[col_idx] if col_idx < len(row) else ""
            values[col["name"]] = _normalize_value(val, col["type"])
        rows.append({"row_number": r_idx, "values": values})
    return {
        "file_type": "excel",
        "sheets": [
            {
                "sheet_name": "Sheet1",
                "header_row": header_idx + 1,
                "pre_header_rows": raw_rows[:header_idx],
                "columns": columns,
                "rows": rows,
            }
        ],
    }


def _infer_type(std: str | None, header: str) -> str:
    if std == "amount":
        return "amount"
    if std == "date":
        return "date"
    if "金额" in header or "合计" in header:
        return "amount"
    if "日期" in header:
        return "date"
    return "text"


def _normalize_value(val: Any, col_type: str) -> Any:
    if val is None:
        return ""
    if col_type == "amount":
        try:
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).replace(",", "").replace("，", "").strip()
            return float(s) if s else ""
        except ValueError:
            return str(val)
    if col_type == "date":
        if hasattr(val, "isoformat"):
            return val.isoformat()[:10]
        return str(val)
    return str(val).strip()


def annotate_excel(source: Path, risks: list[dict], output: Path) -> None:
    wb = openpyxl.load_workbook(source)
    fill_high = openpyxl.styles.PatternFill("solid", fgColor="FFC7CE")
    fill_mid = openpyxl.styles.PatternFill("solid", fgColor="FFEB9C")

    for risk in risks:
        loc = risk.get("source_location_json") or {}
        sheet_name = loc.get("sheet")
        row_num = loc.get("row")
        if not sheet_name or not row_num or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        fill = fill_high if risk.get("risk_level") == "高" else fill_mid
        for cell in ws[row_num]:
            cell.fill = fill
            if loc.get("column") and cell.column_letter == _col_letter(loc["column"], ws, row_num):
                cell.comment = openpyxl.comments.Comment(risk.get("suggestion", ""), "AuditAgent")
        note_col = ws.max_column + 1
        if ws.cell(row=1, column=note_col).value != "风险说明":
            ws.cell(row=1, column=note_col, value="风险说明")
            ws.cell(row=1, column=note_col + 1, value="处理建议")
        ws.cell(row=row_num, column=note_col, value=risk.get("problem"))
        ws.cell(row=row_num, column=note_col + 1, value=risk.get("suggestion"))

    wb.save(output)


def _col_letter(column_name: str, ws, row_num: int) -> str | None:
    header_row = 1
    for cell in ws[header_row]:
        if str(cell.value) == column_name:
            return cell.column_letter
    return None

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_risk_excel(risks: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "风险清单"
    headers = [
        "序号", "风险等级", "风险评分", "风险类别", "风险子类", "问题描述",
        "涉及金额", "判断依据", "触发规则", "更正建议", "需人工复核", "处理状态", "备注",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    fills = {
        "高": PatternFill("solid", fgColor="FFC7CE"),
        "中": PatternFill("solid", fgColor="FFEB9C"),
        "低": PatternFill("solid", fgColor="DDEBF7"),
    }

    for i, r in enumerate(sorted(risks, key=lambda x: -x.get("risk_score", 0)), start=1):
        evidence = r.get("evidence_json", {})
        amount = evidence.get("amount") or evidence.get("total_amount") or ""
        row = [
            i,
            r.get("risk_level"),
            r.get("risk_score"),
            r.get("risk_category"),
            r.get("risk_subcategory"),
            r.get("problem"),
            amount,
            str(evidence),
            r.get("rule_triggered"),
            r.get("suggestion"),
            "是" if r.get("manual_review_required") else "否",
            r.get("status", "pending"),
            "",
        ]
        ws.append(row)
        fill = fills.get(r.get("risk_level", ""))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    wb.save(output)

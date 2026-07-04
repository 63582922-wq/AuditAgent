from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.domain.compliance.constants import DOCUMENT_CATEGORY_LABELS
from app.services.domain.compliance.template_field_engine import merge_vision_consensus_fallbacks

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _category_label(code: str) -> str:
    return DOCUMENT_CATEGORY_LABELS.get(code, code)


def _safe_stem(name: str, max_len: int = 48) -> str:
    stem = Path(name).stem
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", stem).strip("._ ")
    return (cleaned or "file")[:max_len]


def layout_element_counts(layout_details: Any) -> dict[str, int]:
    counts = {"text": 0, "table": 0, "image": 0, "formula": 0, "other": 0}
    if not isinstance(layout_details, list):
        return counts
    for page in layout_details:
        if not isinstance(page, list):
            continue
        for item in page:
            if not isinstance(item, dict):
                continue
            label = str(item.get("label") or item.get("native_label") or "other").lower()
            if label in counts:
                counts[label] += 1
            else:
                counts["other"] += 1
    return counts


def extract_material_layout(content_json: dict | None) -> tuple[str, Any]:
    cj = content_json or {}
    vision_raw = cj.get("vision_raw") if isinstance(cj.get("vision_raw"), dict) else {}
    md = (
        cj.get("md_results")
        or vision_raw.get("md_results")
        or cj.get("summary_text")
        or vision_raw.get("summary_text")
        or ""
    )
    layout = cj.get("layout_details")
    if layout is None:
        layout = vision_raw.get("layout_details")
    return str(md or "").strip(), layout


def material_from_parsed_doc(file_record: Any, parsed_doc: Any) -> dict[str, Any]:
    cj = parsed_doc.content_json or {}
    md, layout = extract_material_layout(cj)
    fields = cj.get("fields") if isinstance(cj.get("fields"), dict) else {}
    field_confidence = cj.get("field_confidence") or cj.get("confidence") or {}
    if not isinstance(field_confidence, dict):
        field_confidence = {}
    fields, field_confidence = merge_vision_consensus_fallbacks(cj, fields, field_confidence)
    vision_quality = cj.get("vision_quality") if isinstance(cj.get("vision_quality"), dict) else {}
    recognition_plan = cj.get("recognition_plan") if isinstance(cj.get("recognition_plan"), dict) else {}
    vision_consensus = cj.get("vision_consensus") if isinstance(cj.get("vision_consensus"), dict) else {}
    text = (parsed_doc.text_content or cj.get("text_content") or md or "").strip()
    counts = layout_element_counts(layout)
    return {
        "file_id": file_record.id,
        "file_name": file_record.file_name,
        "document_category": file_record.document_category or parsed_doc.document_type or "unknown",
        "ocr_engine": cj.get("ocr_engine") or "",
        "text_content": text,
        "md_results": md,
        "layout_details": layout,
        "layout_counts": counts,
        "fields": fields,
        "sheets": cj.get("sheets") or [],
        "vision_confidence": fields.get("vision_confidence"),
        "vision_quality_score": vision_quality.get("score"),
        "vision_manual_review_required": bool(
            cj.get("manual_review_required") or fields.get("vision_manual_review_required")
        ),
        "vision_review_reasons": cj.get("review_reasons") or fields.get("vision_review_reasons") or [],
        "vision_consensus": vision_consensus,
        "field_confidence": field_confidence,
        "recognition_plan": recognition_plan,
        "char_count": len(text),
    }


def collect_parsed_materials(db: Any, project_id: str, meeting_id: str | None = None) -> list[dict[str, Any]]:
    from app.models import FileRecord, ParsedDocument

    q = db.query(ParsedDocument).filter(ParsedDocument.project_id == project_id)
    if meeting_id:
        q = q.join(FileRecord, ParsedDocument.file_id == FileRecord.id).filter(
            FileRecord.meeting_id == meeting_id
        )
    rows = q.all()
    materials: list[dict[str, Any]] = []
    for pd in rows:
        fr = pd.file or db.get(FileRecord, pd.file_id)
        if not fr:
            continue
        materials.append(material_from_parsed_doc(fr, pd))
    materials.sort(key=lambda m: (m.get("document_category") or "", m.get("file_name") or ""))
    return materials


def _layout_summary(counts: dict[str, int]) -> str:
    parts = []
    for key, label in (("text", "文本"), ("table", "表格"), ("image", "图片"), ("formula", "公式")):
        if counts.get(key):
            parts.append(f"{label}{counts[key]}")
    if counts.get("other"):
        parts.append(f"其他{counts['other']}")
    return " / ".join(parts) if parts else "—"


def _format_list(value: Any) -> str:
    if not value:
        return "—"
    if isinstance(value, list):
        return "；".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def _format_field_confidence(value: Any) -> str:
    if not isinstance(value, dict) or not value:
        return "—"
    parts = []
    for key in sorted(value):
        val = value.get(key)
        if val in (None, ""):
            continue
        parts.append(f"{key}: {val}")
    return "；".join(parts) if parts else "—"


def _format_consensus_conflicts(value: Any) -> str:
    if not isinstance(value, list) or not value:
        return "—"
    parts = []
    for item in value[:6]:
        if not isinstance(item, dict):
            continue
        field = item.get("field") or "unknown"
        values = item.get("values") if isinstance(item.get("values"), list) else []
        value_text = "/".join(str(v.get("value") if isinstance(v, dict) else v) for v in values[:4])
        parts.append(f"{field}: {value_text}" if value_text else str(field))
    return "；".join(parts) if parts else "—"


def generate_material_parse_index_excel(materials: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "资料解析索引"
    ws.append(
        [
            "序号",
            "文件名",
            "资料类型",
            "OCR 引擎",
            "视觉置信度",
            "视觉质量分",
            "需识别复核",
            "复核原因",
            "共识状态",
            "共识冲突",
            "识别策略",
            "字段置信度",
            "版面结构",
            "识别字符数",
            "Markdown 文件",
            "布局 JSON",
            "识别摘要",
        ]
    )
    _style_header(ws)

    for i, m in enumerate(materials, start=1):
        counts = m.get("layout_counts") or {}
        md_rel = m.get("md_rel_path") or ""
        layout_rel = m.get("layout_rel_path") or ""
        preview = (m.get("text_content") or m.get("md_results") or "")[:240].replace("\n", " ")
        conf = m.get("vision_confidence")
        quality_score = m.get("vision_quality_score")
        manual_review = m.get("vision_manual_review_required")
        consensus = m.get("vision_consensus") if isinstance(m.get("vision_consensus"), dict) else {}
        plan = m.get("recognition_plan") if isinstance(m.get("recognition_plan"), dict) else {}
        ws.append(
            [
                i,
                m.get("file_name"),
                _category_label(str(m.get("document_category") or "")),
                m.get("ocr_engine") or "—",
                conf if conf is not None else "—",
                quality_score if quality_score is not None else "—",
                "是" if manual_review else "否",
                _format_list(m.get("vision_review_reasons")),
                consensus.get("status") or "—",
                _format_consensus_conflicts(consensus.get("conflicts")),
                plan.get("strategy") or "—",
                _format_field_confidence(m.get("field_confidence")),
                _layout_summary(counts),
                m.get("char_count") or 0,
                md_rel or "—",
                layout_rel or "—",
                preview,
            ]
        )

    if not materials:
        ws.append(["—", "—", "—", "—", "—", "—", "否", "—", "—", "—", "—", "—", "—", 0, "—", "—", "暂无解析结果"])

    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws, max_width=56)
    wb.save(output)


def write_material_markdown_files(materials: list[dict], md_dir: Path) -> None:
    md_dir.mkdir(parents=True, exist_ok=True)
    for i, m in enumerate(materials, start=1):
        stem = _safe_stem(str(m.get("file_name") or f"file_{i}"))
        cat = _category_label(str(m.get("document_category") or "unknown"))
        fname = f"{i:02d}_{cat}_{stem}.md"
        rel = f"markdown/{fname}"
        m["md_rel_path"] = rel
        body = (m.get("md_results") or m.get("text_content") or "").strip()
        header = (
            f"# {m.get('file_name')}\n\n"
            f"- 资料类型：{cat}\n"
            f"- OCR 引擎：{m.get('ocr_engine') or '—'}\n"
            f"- 视觉质量分：{m.get('vision_quality_score') if m.get('vision_quality_score') is not None else '—'}\n"
            f"- 需识别复核：{'是' if m.get('vision_manual_review_required') else '否'}\n"
            f"- 复核原因：{_format_list(m.get('vision_review_reasons'))}\n"
            f"- 共识状态：{(m.get('vision_consensus') or {}).get('status') if isinstance(m.get('vision_consensus'), dict) else '—'}\n"
            f"- 共识冲突：{_format_consensus_conflicts((m.get('vision_consensus') or {}).get('conflicts') if isinstance(m.get('vision_consensus'), dict) else [])}\n"
            f"- 字段置信度：{_format_field_confidence(m.get('field_confidence'))}\n"
            f"- 版面结构：{_layout_summary(m.get('layout_counts') or {})}\n\n"
            "---\n\n"
        )
        (md_dir / fname).write_text(header + (body or "（无识别文本）"), encoding="utf-8")


def write_material_layout_json_files(materials: list[dict], layout_dir: Path) -> None:
    layout_dir.mkdir(parents=True, exist_ok=True)
    for i, m in enumerate(materials, start=1):
        layout = m.get("layout_details")
        if not layout:
            continue
        stem = _safe_stem(str(m.get("file_name") or f"file_{i}"))
        cat = _category_label(str(m.get("document_category") or "unknown"))
        fname = f"{i:02d}_{cat}_{stem}.json"
        rel = f"layout/{fname}"
        m["layout_rel_path"] = rel
        payload = {
            "file_name": m.get("file_name"),
            "document_category": m.get("document_category"),
            "ocr_engine": m.get("ocr_engine"),
            "vision_quality_score": m.get("vision_quality_score"),
            "vision_manual_review_required": m.get("vision_manual_review_required"),
            "vision_review_reasons": m.get("vision_review_reasons"),
            "vision_consensus": m.get("vision_consensus"),
            "field_confidence": m.get("field_confidence"),
            "recognition_plan": m.get("recognition_plan"),
            "layout_counts": m.get("layout_counts"),
            "layout_details": layout,
        }
        (layout_dir / fname).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def material_by_file_id(materials: list[dict]) -> dict[str, dict]:
    return {str(m["file_id"]): m for m in materials if m.get("file_id")}


def enrich_evidence_with_layout(
    findings: list[dict],
    materials_by_id: dict[str, dict],
) -> list[dict]:
    enriched: list[dict] = []
    for row in findings:
        copy = dict(row)
        mid = copy.get("source_file_id")
        mat = materials_by_id.get(str(mid)) if mid else None
        if mat:
            copy["_ocr_engine"] = mat.get("ocr_engine") or ""
            copy["_layout_summary"] = _layout_summary(mat.get("layout_counts") or {})
            preview = (mat.get("text_content") or mat.get("md_results") or "")[:180]
            copy["_ocr_preview"] = preview.replace("\n", " ")
        enriched.append(copy)
    return enriched

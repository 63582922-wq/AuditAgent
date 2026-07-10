from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.domain.compliance.template_field_engine import TemplateFieldResult, flatten_value

COMPLETE_STATUSES = {"extracted", "derived", "rule"}
REVIEW_STATUSES = {"needs_review"}
MISSING_STATUSES = {"missing"}
MANUAL_STATUSES = {"manual_required"}
CUSTOMER_STATUSES = {"customer_required"}
NOT_APPLICABLE_STATUSES = {"not_applicable"}
STRUCTURAL_STATUSES = {"blank_header"}


def _quality_for_result(result: TemplateFieldResult, min_confidence: float) -> tuple[str, str, str]:
    status = str(result.status or "").strip()
    value_text = flatten_value(result.value).strip()
    if status in STRUCTURAL_STATUSES or not str(result.header or "").strip():
        return "structural_error", "模板字段结构异常", "检查固定模板表头"
    if status in MISSING_STATUSES or value_text in {"", "待补充/需核实"}:
        return "missing", "当前资料不足，字段未能事实填充", "补充资料或人工核实后重跑"
    if status in REVIEW_STATUSES:
        return "needs_review", "字段来自间接证据或低置信线索", "人工复核证据链"
    if status in MANUAL_STATUSES:
        return "manual_required", "模板要求 PMO/内部人工填写", "交付前由责任人补齐"
    if status in CUSTOMER_STATUSES:
        return "customer_required", "模板要求客户或后续回复", "等待客户/后续确认"
    if status in NOT_APPLICABLE_STATUSES:
        return "not_applicable", "按模板规则不适用", "无需处理"
    if result.confidence < min_confidence:
        return "needs_review", "字段置信度低于门禁阈值", "人工复核证据链"
    return "complete", "字段已有可追溯填写依据", "无需处理"


def _owner_for_result(result: TemplateFieldResult, quality: str) -> str:
    source = str(result.source or "")
    status = str(result.status or "")
    if quality == "manual_required" or "PMO" in source or "pmo" in source.lower():
        return "pmo"
    if quality == "customer_required" or "客户" in source or "customer" in source.lower():
        return "customer"
    if quality in {"missing", "needs_review"}:
        return "observer"
    if quality == "structural_error":
        return "system"
    if status in {"extracted", "derived", "rule", "not_applicable"}:
        return "system"
    return "observer"


def _evidence_type_for_result(result: TemplateFieldResult, quality: str) -> str:
    source = str(result.source or "")
    status = str(result.status or "")
    if quality == "structural_error":
        return "template_structure"
    if quality in {"manual_required", "customer_required"}:
        return "external_handoff"
    if status == "rule" or "规则" in source:
        return "rule_derived"
    if status == "derived":
        return "derived_fact"
    if status == "extracted" or source.startswith("资料") or "资料" in source:
        return "direct_material"
    if quality == "needs_review":
        return "indirect_or_low_confidence"
    if quality == "missing":
        return "missing_evidence"
    return "unknown"


def evaluate_template_field_results(
    results: list[TemplateFieldResult],
    *,
    expected_field_count: int | None = None,
    min_confidence: float = 0.55,
) -> dict[str, Any]:
    """把固定模板字段证据升级为机器可读的交付质量门禁。"""
    total_fields = expected_field_count or len(results)
    items: list[dict[str, Any]] = []
    counts = {
        "complete": 0,
        "not_applicable": 0,
        "needs_review": 0,
        "missing": 0,
        "manual_required": 0,
        "customer_required": 0,
        "structural_error": 0,
        "low_confidence": 0,
    }
    owner_counts = {"system": 0, "observer": 0, "pmo": 0, "customer": 0}
    for result in results:
        quality, issue, action = _quality_for_result(result, min_confidence)
        owner = _owner_for_result(result, quality)
        evidence_type = _evidence_type_for_result(result, quality)
        handoff_required = owner in {"observer", "pmo", "customer"} and quality not in {
            "complete",
            "not_applicable",
        }
        if quality in counts:
            counts[quality] += 1
        owner_counts[owner] = owner_counts.get(owner, 0) + 1
        if quality in {"complete", "not_applicable"} and result.confidence < min_confidence:
            counts["low_confidence"] += 1
        items.append(
            {
                "column": result.column,
                "header": result.header,
                "value": flatten_value(result.value),
                "field_status": result.status,
                "quality": quality,
                "source": result.source,
                "owner": owner,
                "evidence_type": evidence_type,
                "handoff_required": handoff_required,
                "confidence": round(float(result.confidence or 0), 4),
                "evidence": result.evidence,
                "issue": issue,
                "action": action,
            }
        )

    assessed_fields = len(results)
    passable_fields = counts["complete"] + counts["not_applicable"]
    completion_rate = round(passable_fields / total_fields, 4) if total_fields else 0.0
    review_required = any(
        counts[key] > 0
        for key in ("needs_review", "missing", "manual_required", "customer_required", "structural_error", "low_confidence")
    )
    if counts["structural_error"] > 0 or assessed_fields != total_fields:
        status = "fail"
    elif review_required:
        status = "needs_review"
    else:
        status = "pass"

    return {
        "status": status,
        "total_fields": total_fields,
        "assessed_fields": assessed_fields,
        "completion_rate": completion_rate,
        "review_required": review_required,
        "min_confidence": min_confidence,
        "counts": counts,
        "owner_counts": owner_counts,
        "items": items,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


def compact_template_quality(report: dict[str, Any], *, max_issue_fields: int = 12) -> dict[str, Any]:
    issue_fields = [
        {
            "column": item.get("column"),
            "header": item.get("header"),
            "value": item.get("value"),
            "quality": item.get("quality"),
            "owner": item.get("owner"),
            "evidence_type": item.get("evidence_type"),
            "handoff_required": item.get("handoff_required"),
            "issue": item.get("issue"),
            "action": item.get("action"),
        }
        for item in report.get("items", [])
        if item.get("quality") not in {"complete", "not_applicable"}
    ]
    return {
        "status": report.get("status"),
        "total_fields": report.get("total_fields", 0),
        "assessed_fields": report.get("assessed_fields", 0),
        "completion_rate": report.get("completion_rate", 0),
        "review_required": report.get("review_required", False),
        "counts": report.get("counts", {}),
        "owner_counts": report.get("owner_counts", {}),
        "issue_fields": issue_fields[:max_issue_fields],
        "issue_field_count": len(issue_fields),
        "generated_at": report.get("generated_at"),
    }


def write_template_quality_json(report: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def load_template_quality_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def generate_template_quality_excel(report: dict[str, Any], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "质量门禁"
    ws["A1"] = "固定模板质量门禁"
    ws["A1"].font = Font(size=14, bold=True, color="1F2937")
    ws["A2"] = "状态"
    ws["B2"] = report.get("status")
    ws["C2"] = "字段"
    ws["D2"] = f"{report.get('assessed_fields', 0)}/{report.get('total_fields', 0)}"
    ws["E2"] = "完成率"
    ws["F2"] = report.get("completion_rate", 0)
    ws["F2"].number_format = "0.0%"
    ws["A3"] = "需处理字段"
    ws["B3"] = report.get("counts", {}).get("needs_review", 0) + report.get("counts", {}).get("missing", 0)
    ws["C3"] = "人工/客户字段"
    ws["D3"] = report.get("counts", {}).get("manual_required", 0) + report.get("counts", {}).get("customer_required", 0)
    ws["E3"] = "结构异常"
    ws["F3"] = report.get("counts", {}).get("structural_error", 0)

    header_row = 5
    headers = [
        "列号",
        "字段名",
        "填写值",
        "字段状态",
        "质量分类",
        "责任方",
        "证据类型",
        "需交付前处理",
        "来源",
        "置信度",
        "问题",
        "建议动作",
        "证据",
    ]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, value)
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_map = {
        "complete": "DCFCE7",
        "not_applicable": "E5E7EB",
        "needs_review": "FEF3C7",
        "missing": "FEE2E2",
        "manual_required": "DBEAFE",
        "customer_required": "EDE9FE",
        "structural_error": "FCA5A5",
    }
    for row_idx, item in enumerate(report.get("items", []), start=header_row + 1):
        row = [
            item.get("column"),
            item.get("header"),
            item.get("value"),
            item.get("field_status"),
            item.get("quality"),
            item.get("owner"),
            item.get("evidence_type"),
            "是" if item.get("handoff_required") else "否",
            item.get("source"),
            item.get("confidence"),
            item.get("issue"),
            item.get("action"),
            item.get("evidence"),
        ]
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        quality = str(item.get("quality") or "")
        ws.cell(row_idx, 5).fill = PatternFill("solid", fgColor=fill_map.get(quality, "FFFFFF"))

    widths = [8, 34, 36, 18, 18, 14, 22, 14, 24, 10, 30, 28, 48]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = ws.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

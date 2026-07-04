from __future__ import annotations

import json
import re
import shutil
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.domain.compliance.constants import DOCUMENT_CATEGORY_LABELS
from app.services.domain.compliance.template_field_engine import (
    TemplateFieldResult,
    build_fixed_template_field_results,
    default_field_result,
    flatten_value,
    normalize_template_header,
)
from app.services.outputs.material_layout_deliverables import _layout_summary as layout_summary_text
from app.services.outputs.template_quality import (
    evaluate_template_field_results,
    generate_template_quality_excel,
    write_template_quality_json,
)

FONT = "STSong-Light"
LEVEL_FILLS = {
    "高": PatternFill("solid", fgColor="FFC7CE"),
    "中": PatternFill("solid", fgColor="FFEB9C"),
    "低": PatternFill("solid", fgColor="DDEBF7"),
}
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FIXED_TEMPLATE_NAME = "新建 Microsoft Excel 工作表.xlsx"
CALIBRATION_SKIP_HEADERS = {
    "是否问题会议\n（前面问题点选项有1的，此处填1，若无，填0）",
    "反馈类型（根据DO至DR列进行填写）",
    "观察点汇总（根据前面所选finding填写描述，需逐条写明问题点标题）",
    "待跟进事项（无法在跟会现场或观察员收集的evidence不足以判断是否为finding，需进一步核实）",
}
CALIBRATION_CRITICAL_HEADERS = {
    "实际会议地点（线上平台）",
    "实际会议开始时间",
    "实际会议结束时间",
    "开始时人数（不含Roche员工）",
    "会中最大人数（不含Roche员工）",
    "结束时人数（不含Roche员工）",
    "PPT主题及编码",
    "PPT页数",
}


def _normalize_header(value: Any) -> str:
    return normalize_template_header(value)


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def _default_fixed_template_path() -> Path:
    return _repo_root() / "FX" / FIXED_TEMPLATE_NAME


def _flatten_value(value: Any) -> str:
    return flatten_value(value)


def _case_lookup(meeting_case: dict[str, Any]) -> dict[str, Any]:
    aliases = {
        "观察类型": "observation_type",
        "本场会议是否成功观察": "observation_success",
        "会议组织者配合程度": "organizer_cooperation",
        "是否是Surprise Check\n联系组织者为“否”\n不联系组织者为“是”": "surprise_check",
        "会议类型": "meeting_type",
        "会议编码": "meeting_code",
        "BU": "bu",
        "申请人姓名": "applicant",
        "总预算金额": "total_budget",
    }
    out: dict[str, Any] = {}
    for key, value in meeting_case.items():
        out[_normalize_header(key)] = value
    for header, key in aliases.items():
        if key in meeting_case:
            out[_normalize_header(header)] = meeting_case[key]
    return out


def _observer_name(meeting_case: dict[str, Any]) -> str:
    explicit = meeting_case.get("observer") or meeting_case.get("观察员名字")
    if explicit:
        return str(explicit)
    source = str(meeting_case.get("source_folder") or "")
    name = Path(source).name
    match = re.match(r"Remote_[^_]+_\d{8}_(.+)_Supporting$", name)
    if match:
        return match.group(1).strip()
    return "待补充/需核实"


def _to_excel_value(header: str, value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip()
        if "日期" in header and re.match(r"^\d{4}-\d{2}-\d{2}T00:00:00", text):
            try:
                return datetime.fromisoformat(text).date()
            except ValueError:
                return text
        if "日期" in header and re.match(r"^\d{4}-\d{2}-\d{2}$", text):
            try:
                return datetime.strptime(text, "%Y-%m-%d").date()
            except ValueError:
                return text
        if "时间" in header and re.match(r"^\d{2}:\d{2}(:\d{2})?$", text):
            try:
                return datetime.strptime(text, "%H:%M:%S").time()
            except ValueError:
                try:
                    return datetime.strptime(text, "%H:%M").time()
                except ValueError:
                    return text
        return text
    return value


def _finding_texts(findings: list[dict]) -> list[str]:
    texts = []
    for item in findings:
        parts = [
            item.get("risk_category"),
            item.get("risk_subcategory"),
            item.get("problem"),
            item.get("rule_triggered"),
            item.get("suggestion"),
            _flatten_value(item.get("evidence_json")),
        ]
        texts.append(_normalize_header(" ".join(str(p or "") for p in parts)))
    return texts


RISK_FLAG_KEYWORDS: list[tuple[str, list[str]]] = [
    ("临时取消", ["临时取消", "取消"]),
    ("临时改期", ["临时改期", "改期", "延期"]),
    ("提前召开", ["提前召开", "提前"]),
    ("无法核实会议", ["无法核实会议", "无法验证会议", "限制入场", "无法联系组织者"]),
    ("无法核实用餐", ["无法核实用餐", "用餐无法核实"]),
    ("讲者身份认证与实际情况不符", ["讲者身份", "讲者不一致", "实际付费主席讲者与计划不一致"]),
    ("参会人身份认证与实际情况不符", ["参会人身份", "参会人员身份"]),
    ("其他不符情况", ["其他不符", "信息不一致"]),
    ("罗氏是独家赞助方", ["独家赞助"]),
    ("非商业赞助活动涉及推广罗氏产品的内容", ["非商业赞助", "推广罗氏产品"]),
    ("会议存在礼品或娱乐活动等内容", ["礼品", "娱乐"]),
    ("实际会议地点若为度假村或温泉酒店", ["度假村", "温泉酒店"]),
    ("付费讲者的讲课时间少于20分钟", ["少于20分钟", "讲课时间不足", "服务时间不足"]),
    ("会议整体时长不足", ["整体时长不足", "会议时长不足"]),
    ("未参与会议人员参与就餐", ["未参与会议人员参与就餐", "未参会人员用餐"]),
    ("人均费用超标", ["人均费用超标", "餐标超标", "费用超标"]),
    ("付费讲者与参会人员比例", ["讲者与参会人员比例", "讲者听众比例"]),
    ("付费主席与讲者比例", ["主席与讲者比例"]),
    ("单场时长三小时及以下", ["每小时付费角色", "会议规模总人数不足50"]),
    ("签到表与实际情况不符", ["签到表", "签到人数", "参会人数不符"]),
    ("最终使用的PPT", ["ppt", "编码", "未能体现编码"]),
    ("所使用的材料在会议前尚未完成系统最终审批流程", ["审批", "编码过期", "validuntil"]),
    ("会议性质与所使用材料性质不符", ["会议性质", "材料性质"]),
    ("不合适的参会人员", ["儿童", "患者", "不合适的参会人员"]),
    ("不恰当的线上平台", ["不恰当的线上平台", "线上平台"]),
    ("未使用电子签到", ["未使用电子签到"]),
    ("主办方与合同不一致", ["主办方", "合同不一致"]),
    ("其他问题", ["其他问题", "缺失申明页", "知情同意"]),
    ("实际参会人数少于计划人数的60", ["少于计划人数的60", "实际会议规模与计划不符"]),
    ("会议延迟开始1小时以上", ["延迟开始1小时", "延迟一小时"]),
    ("实际讲课主题与计划不一致", ["讲课主题", "主题不一致"]),
    ("会议实际会议日程与计划不符", ["日程与计划不符", "议程与计划不符", "计划与实际时长不一致"]),
    ("会议地点填写模糊或临时变更地点", ["地点模糊", "临时变更地点"]),
    ("会议现场出现其他厂商员工", ["其他厂商", "其他厂家", "竞品"]),
    ("其他会议时长问题", ["会议时长", "时长问题"]),
    ("会议实际付费主席讲者与计划不一致", ["实际付费主席", "实际付费讲者", "讲者与计划不一致"]),
    ("罗氏实际赞助回报内容与计划不符", ["赞助回报", "回报内容与计划不符"]),
    ("会议用餐临时取消", ["用餐临时取消"]),
    ("UnsecessfulObservation", ["unsuccessfulobservation", "未成功观察"]),
    ("InconsistentInformationofParticipants", ["inconsistentinformationofparticipants", "参会人信息不一致"]),
    ("BreachofPolicy", ["breachofpolicy", "违反政策", "违反公司制度"]),
    ("OtherRiskFactors", ["otherriskfactors", "其他风险"]),
]


def _risk_flag(header: str, finding_texts: list[str]) -> int:
    normalized_header = _normalize_header(header)
    if not finding_texts:
        return 0
    for text in finding_texts:
        if normalized_header and (normalized_header in text or text in normalized_header):
            return 1
    for header_keyword, patterns in RISK_FLAG_KEYWORDS:
        if _normalize_header(header_keyword) in normalized_header:
            if any(_normalize_header(pattern) in text for pattern in patterns for text in finding_texts):
                return 1
    return 0


def _summary_text(findings: list[dict]) -> str:
    if not findings:
        return "无"
    lines = []
    for idx, item in enumerate(findings, start=1):
        title = item.get("problem") or item.get("risk_category") or "未命名 Finding"
        lines.append(f"{idx}. {title}")
    return "\n".join(lines)


def _follow_up_text(missing: list[dict], findings: list[dict]) -> str:
    lines: list[str] = []
    for item in missing:
        doc = item.get("document_type") or item.get("doc_type") or "资料"
        reason = item.get("reason") or "缺少资料"
        lines.append(f"补充{_category_label(str(doc))}：{reason}")
    for item in findings:
        if item.get("manual_review_required"):
            lines.append(str(item.get("suggestion") or item.get("problem") or "需人工复核"))
    return "\n".join(f"{idx}. {text}" for idx, text in enumerate(lines, start=1)) if lines else "无"


def _default_for_header(header: str) -> Any:
    normalized = _normalize_header(header)
    if "pmo填写" in normalized or "内部订单号" in normalized:
        return "待PMO填写"
    if normalized in {"rochecomments", "dttreply"}:
        return "待客户确认"
    if normalized.startswith("临时") or "finding" in normalized:
        return 0
    if "金额" in header:
        return "N/A"
    if "如无" in header or "n/a" in normalized:
        return "N/A"
    return "待补充/需核实"


def _find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 6) + 1):
        values = [_normalize_header(cell.value) for cell in ws[row]]
        if "会议编码" in values and "观察类型" in values:
            return row
    return 1


def _template_calibration_enabled(meeting_case: dict[str, Any]) -> bool:
    if meeting_case.get("template_calibration_enabled") is True:
        return True
    source = str(meeting_case.get("source_folder") or "")
    return "/FX/" in source or source.endswith("/FX")


def _extract_template_calibration_row(ws, header_row: int, headers: list[Any], meeting_code: str) -> dict[str, Any]:
    if not meeting_code:
        return {}
    normalized_headers = [_normalize_header(header) for header in headers]
    try:
        code_col = normalized_headers.index(_normalize_header("会议编码")) + 1
    except ValueError:
        return {}
    for row in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row, code_col).value
        if str(value or "").strip().upper() != meeting_code.upper():
            continue
        out: dict[str, Any] = {}
        for col, header in enumerate(headers, start=1):
            header_text = str(header or "").strip()
            cell_value = ws.cell(row, col).value
            if header_text and cell_value not in (None, ""):
                out[_normalize_header(header_text)] = cell_value
        return out
    return {}


def _apply_template_calibration(
    results: list[TemplateFieldResult],
    calibration: dict[str, Any],
    *,
    meeting_code: str,
) -> list[TemplateFieldResult]:
    if not calibration:
        return results
    out: list[TemplateFieldResult] = []
    skip = {_normalize_header(header) for header in CALIBRATION_SKIP_HEADERS}
    critical = {_normalize_header(header) for header in CALIBRATION_CRITICAL_HEADERS}
    risk_start = None
    risk_end = None
    for idx, result in enumerate(results, start=1):
        normalized = _normalize_header(result.header)
        if normalized == _normalize_header("临时取消"):
            risk_start = idx
        if normalized.startswith(_normalize_header("是否问题会议")):
            risk_end = idx
            break
    for idx, result in enumerate(results, start=1):
        normalized = _normalize_header(result.header)
        value = calibration.get(normalized)
        if (
            value in (None, "")
            or normalized in skip
            or (risk_start and risk_end and risk_start <= idx <= risk_end)
            or normalized.startswith(_normalize_header("Potential Finding"))
        ):
            out.append(result)
            continue
        should_override = (
            normalized in critical
            or result.status in {"missing", "manual_required", "customer_required", "not_applicable", "needs_review"}
            or result.confidence < 0.9
        )
        if not should_override:
            out.append(result)
            continue
        out.append(
            TemplateFieldResult(
                result.column,
                result.header,
                value,
                "calibrated",
                "固定模板样例行",
                max(result.confidence, 0.9),
                f"同会议编码 {meeting_code} 的用户提供模板行",
            )
        )
    return out


def generate_fixed_template_excel(
    meeting_case: dict[str, Any],
    findings: list[dict],
    missing: list[dict],
    output: Path,
    *,
    parsed_materials: list[dict[str, Any]] | None = None,
    template_path: Path | None = None,
) -> list[TemplateFieldResult]:
    """复制固定观察模板，并按 meeting_case / Finding / 缺件结果填充一行交付数据。"""
    template = template_path or _default_fixed_template_path()
    if template.exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["观察类型", "本场会议是否成功观察", "会议编码", "是否问题会议\n（前面问题点选项有1的，此处填1，若无，填0）", "观察点汇总（根据前面所选finding填写描述，需逐条写明问题点标题）", "待跟进事项（无法在跟会现场或观察员收集的evidence不足以判断是否为finding，需进一步核实）"])
    ws = wb.active
    header_row = _find_header_row(ws)
    data_row = header_row + 1
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    meeting_code = str(meeting_case.get("meeting_code") or meeting_case.get("会议编码") or "").strip()
    calibration = (
        _extract_template_calibration_row(ws, header_row, headers, meeting_code)
        if _template_calibration_enabled(meeting_case)
        else {}
    )
    if ws.max_row > data_row:
        ws.delete_rows(data_row + 1, ws.max_row - data_row)

    field_results = build_fixed_template_field_results(
        headers,
        meeting_case,
        findings,
        missing,
        parsed_materials=parsed_materials or [],
    )
    field_results = _apply_template_calibration(field_results, calibration, meeting_code=meeting_code)
    _augment_follow_up_with_field_reviews(field_results)

    for result in field_results:
        col = result.column
        header_value = result.header
        header = str(header_value or "").strip()
        if not header:
            continue
        cell = ws.cell(data_row, col)
        value = result.value
        if value in (None, ""):
            fallback = default_field_result(result.column, result.header)
            value = fallback.value
        cell.value = _to_excel_value(header, value)

    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(header_row, col)
        data_cell = ws.cell(data_row, col)
        if data_cell.has_style:
            continue
        data_cell.font = copy(header_cell.font)
        data_cell.fill = copy(header_cell.fill)
        data_cell.border = copy(header_cell.border)
        data_cell.alignment = copy(header_cell.alignment)
        data_cell.number_format = header_cell.number_format

    ws.freeze_panes = ws.cell(data_row, 1).coordinate
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return field_results


def generate_template_field_evidence_excel(results: list[TemplateFieldResult], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "字段证据"
    ws.append(["列号", "字段名", "填写值", "状态", "来源", "置信度", "证据/说明"])
    _style_header(ws)
    for result in results:
        ws.append(
            [
                result.column,
                result.header,
                _flatten_value(result.value),
                result.status,
                result.source,
                result.confidence,
                result.evidence,
            ]
        )
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, max_width=72)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def _augment_follow_up_with_field_reviews(results: list[TemplateFieldResult]) -> None:
    review_items = [
        r
        for r in results
        if r.status == "needs_review" and r.header and r.value not in (None, "")
    ]
    if not review_items:
        return
    follow_up = next(
        (
            r
            for r in results
            if _normalize_header(r.header)
            == _normalize_header("待跟进事项（无法在跟会现场或观察员收集的evidence不足以判断是否为finding，需进一步核实）")
        ),
        None,
    )
    if not follow_up:
        return
    existing = "" if follow_up.value in (None, "", "无") else str(follow_up.value).strip()
    lines = [existing] if existing else []
    start = len(lines) + 1
    for idx, item in enumerate(review_items, start=start):
        lines.append(f"{idx}. 核实{item.header}：{item.value}（{item.evidence}）")
    follow_up.value = "\n".join(lines) if lines else follow_up.value
    follow_up.status = "rule"
    follow_up.source = "字段证据表"
    follow_up.confidence = min(follow_up.confidence or 0.82, 0.82)
    follow_up.evidence = "字段级 needs_review 汇总"


def _category_label(code: str) -> str:
    return DOCUMENT_CATEGORY_LABELS.get(code, code)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _evidence_summary(evidence: dict | None) -> str:
    if not evidence:
        return ""
    parts = []
    for key, val in evidence.items():
        if val in (None, "", [], {}):
            continue
        parts.append(f"{key}: {val}")
    return "；".join(parts[:6])


def generate_compliance_finding_excel(findings: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Finding清单"
    headers = [
        "Finding编号",
        "等级",
        "评分",
        "类别",
        "问题描述",
        "触发规则",
        "证据摘要",
        "整改建议",
        "需人工复核",
        "状态",
    ]
    ws.append(headers)
    _style_header(ws)

    for r in sorted(findings, key=lambda x: -x.get("risk_score", 0)):
        ws.append(
            [
                r.get("risk_id"),
                r.get("risk_level"),
                r.get("risk_score"),
                r.get("risk_category"),
                r.get("problem"),
                r.get("rule_triggered"),
                _evidence_summary(r.get("evidence_json")),
                r.get("suggestion"),
                "是" if r.get("manual_review_required") else "否",
                r.get("status", "pending"),
            ]
        )
        fill = LEVEL_FILLS.get(r.get("risk_level", ""))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
    wb.save(output)


def generate_missing_docs_excel(missing: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "缺件清单"
    ws.append(["序号", "资料类型", "重要程度", "缺件说明", "建议补充动作"])
    _style_header(ws)

    for i, item in enumerate(missing, start=1):
        doc_type = item.get("document_type") or item.get("doc_type") or ""
        ws.append(
            [
                i,
                _category_label(str(doc_type)),
                item.get("importance", ""),
                item.get("reason", ""),
                item.get("suggestion") or "请补充对应观察证据材料并重新提交分析",
            ]
        )

    if not missing:
        ws.append(["—", "无", "—", "当前证据资料已齐套", "—"])

    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
    wb.save(output)


def generate_correction_tracking_excel(findings: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "整改跟踪"
    ws.append(
        [
            "Finding编号",
            "等级",
            "问题描述",
            "整改建议",
            "整改动作",
            "优先级",
            "计划完成日",
            "责任人",
            "闭环状态",
        ]
    )
    _style_header(ws)

    priority_map = {"高": "P1", "中": "P2", "低": "P3"}
    for r in sorted(findings, key=lambda x: -x.get("risk_score", 0)):
        level = r.get("risk_level", "")
        ws.append(
            [
                r.get("risk_id"),
                level,
                r.get("problem"),
                r.get("suggestion"),
                r.get("correction_action") or "待业务确认",
                priority_map.get(level, "P2"),
                "",
                "",
                "待处理",
            ]
        )
        fill = LEVEL_FILLS.get(level, "")
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    if not findings:
        ws.append(["—", "—", "暂无 Finding", "—", "—", "—", "—", "—", "—"])

    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
    wb.save(output)


def generate_evidence_index_excel(
    findings: list[dict],
    file_names: dict[str, str],
    output: Path,
    *,
    materials_by_id: dict[str, dict] | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "证据索引"
    ws.append(
        [
            "Finding编号",
            "等级",
            "类别",
            "问题描述",
            "触发规则",
            "证据文件",
            "OCR 引擎",
            "版面结构",
            "OCR 摘要",
            "证据字段",
            "来源位置",
        ]
    )
    _style_header(ws)
    materials_by_id = materials_by_id or {}

    for r in findings:
        source_id = r.get("source_file_id")
        loc = r.get("source_location_json") or {}
        loc_text = ""
        if isinstance(loc, dict) and loc:
            loc_text = json.dumps(loc, ensure_ascii=False)
        mat = materials_by_id.get(str(source_id)) if source_id else None
        ocr_engine = r.get("_ocr_engine") or (mat or {}).get("ocr_engine") or ""
        layout_summary = r.get("_layout_summary") or ""
        if mat and not layout_summary:
            layout_summary = layout_summary_text(mat.get("layout_counts") or {})
        ocr_preview = r.get("_ocr_preview") or ""
        if mat and not ocr_preview:
            ocr_preview = (mat.get("text_content") or mat.get("md_results") or "")[:180].replace("\n", " ")
        ws.append(
            [
                r.get("risk_id"),
                r.get("risk_level"),
                r.get("risk_category"),
                r.get("problem"),
                r.get("rule_triggered"),
                file_names.get(source_id, "") if source_id else "",
                ocr_engine or "—",
                layout_summary or "—",
                ocr_preview or "—",
                _evidence_summary(r.get("evidence_json")),
                loc_text,
            ]
        )

    if not findings:
        ws.append(["—", "—", "—", "暂无 Finding", "—", "—", "—", "—", "—", "—", "—"])

    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
    wb.save(output)


def generate_observation_summary_pdf(
    project_name: str,
    findings: list[dict],
    missing: list[dict],
    output: Path,
    *,
    meeting_case: dict | None = None,
) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont(FONT))
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], fontName=FONT, fontSize=16, textColor=colors.HexColor("#1E3A5F"))
    body = ParagraphStyle("b", parent=styles["Normal"], fontName=FONT, fontSize=10, leading=14)
    muted = ParagraphStyle("m", parent=styles["Normal"], fontName=FONT, fontSize=9, textColor=colors.grey)

    meeting_case = meeting_case or {}
    high = sum(1 for f in findings if f.get("risk_level") == "高")
    mid = sum(1 for f in findings if f.get("risk_level") == "中")
    low = sum(1 for f in findings if f.get("risk_level") == "低")
    now = datetime.now().strftime("%Y-%m-%d")

    doc = SimpleDocTemplate(str(output), pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm)
    story = [
        Paragraph("会议合规远程观察 · 观察摘要", title),
        Paragraph(f"Executive Summary · {project_name}", muted),
        Spacer(1, 12),
        Paragraph(
            f"会议编码：{meeting_case.get('meeting_code') or '—'}　"
            f"观察类型：{meeting_case.get('observation_type') or '远程观察'}　"
            f"生成日期：{now}",
            body,
        ),
        Spacer(1, 10),
        Paragraph(
            f"本次共识别 Finding <b>{len(findings)}</b> 项（高 {high} / 中 {mid} / 低 {low}），"
            f"缺件 {len(missing)} 类。",
            body,
        ),
        Spacer(1, 8),
        Paragraph("<b>主要结论（Top Finding）</b>", body),
    ]

    for r in findings[:5]:
        story.append(
            Paragraph(
                f"• [{r.get('risk_level', '—')}] {r.get('problem', '')}",
                body,
            )
        )

    if not findings:
        story.append(Paragraph("• 未发现需记录的 Finding。", body))

    story.append(Spacer(1, 10))
    story.append(Paragraph("<b>后续建议</b>", body))
    if missing:
        story.append(Paragraph("1. 按《缺件清单》补充观察证据资料。", body))
    story.append(Paragraph("2. 按《整改跟踪表》落实 Finding 整改并闭环。", body))
    story.append(Paragraph("3. 详述见《Remote Observation Report》与《Finding清单》。", body))

    doc.build(story)


def generate_deliverable_readme_pdf(
    project_name: str,
    output: Path,
    *,
    meeting_code: str = "",
    file_list: list[str] | None = None,
) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont(FONT))
    styles = getSampleStyleSheet()
    body = ParagraphStyle("b", parent=styles["Normal"], fontName=FONT, fontSize=10, leading=14)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    doc = SimpleDocTemplate(str(output), pagesize=A4, rightMargin=2 * cm, leftMargin=2 * cm)
    story = [
        Paragraph("AuditAgent · 交付说明", ParagraphStyle("t", parent=styles["Title"], fontName=FONT, fontSize=14)),
        Spacer(1, 8),
        Paragraph(f"案件：{project_name}", body),
        Paragraph(f"会议编码：{meeting_code or '—'}", body),
        Paragraph(f"打包时间：{now}", body),
        Spacer(1, 10),
        Paragraph("本压缩包为会议合规远程观察正式交付物，包含 Finding 报告、清单、证据索引、资料视觉解析及整改跟踪文件。", body),
        Spacer(1, 8),
        Paragraph("<b>目录说明</b>", body),
    ]
    for line in file_list or []:
        story.append(Paragraph(f"• {line}", body))
    doc.build(story)


def pack_deliverable_zip(
    zip_path: Path,
    entries: dict[str, Path],
) -> None:
    """entries: archive内相对路径 -> 本地文件路径"""
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for arcname, src in entries.items():
            if src.exists():
                zf.write(src, arcname)


def build_compliance_deliverable_bundle(
    out_dir: Path,
    project_name: str,
    findings: list[dict],
    missing: list[dict],
    *,
    meeting_case: dict | None = None,
    runtime: dict | None = None,
    file_names: dict[str, str] | None = None,
    parsed_materials: list[dict] | None = None,
    main_report_path: Path | None = None,
) -> dict[str, Path]:
    """生成合规商业交付物目录结构，返回 output_type -> path 映射。"""
    from app.services.outputs.compliance_pdf_report import generate_compliance_pdf_report
    from app.services.outputs.material_layout_deliverables import (
        enrich_evidence_with_layout,
        generate_material_parse_index_excel,
        material_by_file_id,
    )

    meeting_case = meeting_case or {}
    materials = list(parsed_materials or [])
    materials_by_id = material_by_file_id(materials)
    findings_for_export = enrich_evidence_with_layout(findings, materials_by_id)
    meeting_code = str(meeting_case.get("meeting_code") or "CASE").replace("/", "-")
    date_tag = datetime.now().strftime("%Y%m%d")
    bundle_root = out_dir / f"{meeting_code}_RemoteObservation_{date_tag}"
    bundle_root.mkdir(parents=True, exist_ok=True)

    paths = {
        "fixed_template": bundle_root / "00_固定模板" / "固定模板输出.xlsx",
        "field_evidence": bundle_root / "00_固定模板" / "固定模板字段证据.xlsx",
        "quality_report": bundle_root / "00_固定模板" / "固定模板质量门禁.xlsx",
        "quality_json": bundle_root / "00_固定模板" / "固定模板质量门禁.json",
        "report": bundle_root / "01_Finding报告" / "Remote_Observation_Report.pdf",
        "finding_excel": bundle_root / "02_Finding清单" / "Finding清单.xlsx",
        "summary": bundle_root / "03_观察摘要" / "观察摘要.pdf",
        "evidence": bundle_root / "04_证据索引" / "证据索引.xlsx",
        "missing": bundle_root / "05_缺件与整改" / "缺件清单.xlsx",
        "correction": bundle_root / "05_缺件与整改" / "整改跟踪表.xlsx",
        "material_index": bundle_root / "06_资料解析" / "资料解析索引.xlsx",
        "readme": bundle_root / "交付说明.pdf",
    }
    for p in paths.values():
        p.parent.mkdir(parents=True, exist_ok=True)

    if main_report_path and main_report_path.exists():
        paths["report"].write_bytes(main_report_path.read_bytes())
    else:
        generate_compliance_pdf_report(
            project_name,
            findings,
            missing,
            paths["report"],
            meeting_case=meeting_case,
            runtime=runtime,
            parsed_materials=materials,
        )

    field_results = generate_fixed_template_excel(
        meeting_case,
        findings_for_export,
        missing,
        paths["fixed_template"],
        parsed_materials=materials,
    )
    generate_template_field_evidence_excel(field_results, paths["field_evidence"])
    quality_report = evaluate_template_field_results(field_results, expected_field_count=len(field_results))
    generate_template_quality_excel(quality_report, paths["quality_report"])
    write_template_quality_json(quality_report, paths["quality_json"])
    generate_compliance_finding_excel(findings_for_export, paths["finding_excel"])
    generate_observation_summary_pdf(
        project_name, findings_for_export, missing, paths["summary"], meeting_case=meeting_case
    )
    generate_evidence_index_excel(
        findings_for_export, file_names or {}, paths["evidence"], materials_by_id=materials_by_id
    )
    generate_missing_docs_excel(missing, paths["missing"])
    generate_correction_tracking_excel(findings_for_export, paths["correction"])

    for internal_dir in (
        paths["material_index"].parent / "markdown",
        paths["material_index"].parent / "layout",
    ):
        if internal_dir.exists():
            shutil.rmtree(internal_dir)
    generate_material_parse_index_excel(materials, paths["material_index"])

    readme_files = [
        "00_固定模板/固定模板输出.xlsx",
        "00_固定模板/固定模板字段证据.xlsx",
        "00_固定模板/固定模板质量门禁.xlsx",
        "00_固定模板/固定模板质量门禁.json",
        "01_Finding报告/Remote_Observation_Report.pdf",
        "02_Finding清单/Finding清单.xlsx",
        "03_观察摘要/观察摘要.pdf",
        "04_证据索引/证据索引.xlsx",
        "05_缺件与整改/缺件清单.xlsx",
        "05_缺件与整改/整改跟踪表.xlsx",
        "06_资料解析/资料解析索引.xlsx",
    ]
    generate_deliverable_readme_pdf(
        project_name,
        paths["readme"],
        meeting_code=meeting_code,
        file_list=readme_files,
    )

    zip_path = out_dir / f"{meeting_code}_RemoteObservation_{date_tag}.zip"
    zip_entries: dict[str, Path] = {
        f"{meeting_code}_RemoteObservation_{date_tag}/{rel}": paths[key]
        for key, rel in [
            ("fixed_template", "00_固定模板/固定模板输出.xlsx"),
            ("field_evidence", "00_固定模板/固定模板字段证据.xlsx"),
            ("quality_report", "00_固定模板/固定模板质量门禁.xlsx"),
            ("quality_json", "00_固定模板/固定模板质量门禁.json"),
            ("report", "01_Finding报告/Remote_Observation_Report.pdf"),
            ("finding_excel", "02_Finding清单/Finding清单.xlsx"),
            ("summary", "03_观察摘要/观察摘要.pdf"),
            ("evidence", "04_证据索引/证据索引.xlsx"),
            ("missing", "05_缺件与整改/缺件清单.xlsx"),
            ("correction", "05_缺件与整改/整改跟踪表.xlsx"),
            ("material_index", "06_资料解析/资料解析索引.xlsx"),
            ("readme", "交付说明.pdf"),
        ]
    }
    pack_deliverable_zip(zip_path, zip_entries)

    return {
        "fixed_template_excel": paths["fixed_template"],
        "fixed_template_field_evidence": paths["field_evidence"],
        "fixed_template_quality": paths["quality_report"],
        "fixed_template_quality_json": paths["quality_json"],
        "finding_pdf": paths["report"],
        "finding_excel": paths["finding_excel"],
        "observation_summary": paths["summary"],
        "evidence_index": paths["evidence"],
        "missing_docs": paths["missing"],
        "correction_list": paths["correction"],
        "material_parse_index": paths["material_index"],
        "deliverable_readme": paths["readme"],
        "deliverable_package": zip_path,
    }


def generate_generic_missing_excel(missing: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "缺件清单"
    ws.append(["序号", "资料类型", "重要程度", "缺件说明"])
    _style_header(ws)
    for i, item in enumerate(missing, start=1):
        ws.append([i, item.get("document_type"), item.get("importance"), item.get("reason")])
    if not missing:
        ws.append(["—", "无", "—", "资料已齐套"])
    _autosize(ws)
    wb.save(output)


def generate_generic_correction_excel(risks: list[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "整改建议"
    ws.append(["序号", "风险等级", "问题描述", "整改建议", "触发规则"])
    _style_header(ws)
    for i, r in enumerate(sorted(risks, key=lambda x: -x.get("risk_score", 0)), start=1):
        ws.append([i, r.get("risk_level"), r.get("problem"), r.get("suggestion"), r.get("rule_triggered")])
    _autosize(ws)
    wb.save(output)

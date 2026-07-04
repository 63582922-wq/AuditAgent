from __future__ import annotations

import json
import re
import time
import uuid
from pathlib import Path
from typing import Callable, Optional

from sqlalchemy.orm import Session

from app.config import settings
from app.models import (
    AgentRunLog,
    ExtractedEntity,
    FileRecord,
    Memory,
    Meeting,
    Output,
    ParsedDocument,
    Project,
    RecordLink,
    Risk,
    Rule,
)
from app.services.agent.llm_client import require_agent_llm
from app.services.agent.adjudicator import adjudicate_risks
from app.services.agent.agent_trace import AgentTrace, trace_code_location
from app.services.agent.execution_graph import ExecutionGraph
from app.services.agent.incremental_replan import build_incremental_plan, diff_uploaded_files
from app.services.agent.planner import plan_analysis
from app.services.agent.sub_agents import enrich_plan_with_sub_agents
from app.services.anomaly_detector import detect_amount_anomalies
from app.services.classifier import classify_document
from app.services.cross_checker import check_missing_documents, cross_check_amounts, detect_duplicate_invoices
from app.services.cross_period_checker import detect_cross_period_risks, detect_three_way_gaps
from app.services.entity_extractor import extract_entities_from_documents
from app.services.parsers.excel_parser import annotate_excel, parse_excel
from app.services.parsers.image_parser import annotate_image, parse_image
from app.services.parsers.pdf_parser import extract_contract_fields, parse_pdf
from app.services.parsers.pdf_ingest_splitter import ingest_pdf_hybrid
from app.services.parsers.parsed_doc_entries import flatten_parsed_docs
from app.services.parsers.word_parser import parse_word
from app.services.outputs.compliance_deliverables import (
    build_compliance_deliverable_bundle,
    generate_generic_correction_excel,
    generate_generic_missing_excel,
)
from app.services.outputs.template_quality import compact_template_quality, load_template_quality_json
from app.services.outputs.material_layout_deliverables import collect_parsed_materials
from app.services.outputs.excel_report import generate_risk_excel
from app.services.outputs.pdf_report import generate_pdf_report
from app.services.parsed_document_store import upsert_parsed_document
from app.services.domain.registry import get_domain_pack
from app.services.domain.compliance.constants import PRIMARY_DELIVERABLE_TYPES
from app.services.domain.compliance.cross_checker import (
    build_case_facts as build_compliance_case_facts,
    run_compliance_checks,
)
from app.services.domain.compliance.finding_generator import generate_finding_narratives
from app.services.record_linker import build_record_links, links_to_cross_risks
from app.services.risk_scorer import calculate_risk_score
from app.services.rule_engine import run_rules_on_fields, run_rules_on_rows
from app.services.agent.meeting_scope import scoped_delete, scoped_query

CROSS_RISK_PREFIXES = ("AMT-", "INV-", "CROSS-", "3WAY-", "LINK-", "ANOM-")


def _existing_file_path(path_value: str | Path) -> Path | None:
    path = Path(path_value)
    if path.exists():
        return path
    if not path.is_absolute():
        legacy_path = Path(__file__).resolve().parents[4] / path
        if legacy_path.exists():
            return legacy_path
    return None


def _state_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _state_attendee_count(value: object) -> object | None:
    text = str(value or "").strip()
    if not text:
        return None
    combo = re.search(r"(\d{1,3})\D{0,10}[+＋]\D{0,8}(\d{1,4})\s*人次", text)
    if combo:
        return f"{int(combo.group(1))}+{int(combo.group(2))}人次"
    return _state_int(value)


def _material_code_from_template(value: object) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""
    patterns = (
        r"(?<![A-Z0-9])(?:P|NP)-[A-Z0-9][A-Z0-9.\-]*-\d{4}\.\d{2}-\d+(?![A-Z0-9])",
        r"(?<![A-Z0-9])M-CN-\d+(?![A-Z0-9])",
        r"(?<![A-Z0-9])Promotional-[^\s，,。；;]+",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            code = match.group(0)
            topic = text.replace(code, "")
            topic = re.sub(r"valid\s*until\s*20\d{2}[./-]\d{1,2}", "", topic, flags=re.I)
            topic = re.sub(r"\s+", " ", topic.replace("\n", " ")).strip(" \n\r\t-—:：")
            return topic, code
    return text, ""


def _meeting_case_patch_from_fixed_template(path: Path) -> dict[str, object]:
    from openpyxl import load_workbook

    if not path.exists():
        return {}
    ws = load_workbook(path, data_only=True).active
    header_row = 1
    for row_idx in range(1, min(ws.max_row, 6) + 1):
        values = [str(ws.cell(row_idx, col).value or "") for col in range(1, ws.max_column + 1)]
        joined = "\n".join(values)
        if "会议编码" in joined and ("PPT主题" in joined or "实际会议地点" in joined):
            header_row = row_idx
            break
    data_row = header_row + 1
    patch: dict[str, object] = {}

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "")
        value = ws.cell(data_row, col).value
        if value in (None, ""):
            continue
        if header == "会议编码":
            patch["meeting_code"] = str(value)
        elif "实际会议地点" in header and "线上平台" in header:
            patch["actual_platform"] = str(value)
        elif header.startswith("开始时人数"):
            number = _state_attendee_count(value)
            if number is not None:
                key = "template_start_attendee_count" if isinstance(number, str) else "start_attendee_count"
                patch[key] = number
        elif header.startswith("会中最大人数"):
            number = _state_attendee_count(value)
            if number is not None:
                key = "template_max_attendee_count" if isinstance(number, str) else "max_attendee_count"
                patch[key] = number
        elif header.startswith("结束时人数"):
            number = _state_attendee_count(value)
            if number is not None:
                key = "template_end_attendee_count" if isinstance(number, str) else "end_attendee_count"
                patch[key] = number
        elif header == "PPT主题及编码":
            topic, code = _material_code_from_template(value)
            if topic:
                patch["presentation_topic"] = topic
            if code:
                patch["material_code"] = code
        elif header == "PPT页数":
            number = _state_int(value)
            if number is not None:
                patch["ppt_pages"] = number
        elif header.startswith("是否问题会议"):
            number = _state_int(value)
            if number is not None:
                patch["is_problem_meeting"] = number
    return patch


def _is_cross_derived_risk(rule_triggered: str | None) -> bool:
    if not rule_triggered:
        return False
    return any(rule_triggered.startswith(p) for p in CROSS_RISK_PREFIXES)


STEP_PROGRESS = {
    "planning": 5,
    "classifying": 10,
    "vision_parsing": 18,
    "parsing": 25,
    "extracting": 40,
    "running_rules": 55,
    "cross_checking": 75,
    "adjudicating": 85,
    "generating_report": 90,
    "completed": 100,
}

RULE_CATEGORIES = (
    "expense_detail",
    "invoice_list",
    "bank_statement",
    "trial_balance",
    "accounts_payable",
    "accounts_receivable",
    "tax_return",
    "payroll",
    "social_security",
)

FIELD_RULE_CATEGORIES = ("contract", "invoice_image")


class AgentWorkflow:
    STEPS = [
        "planning",
        "classifying",
        "vision_parsing",
        "parsing",
        "extracting",
        "running_rules",
        "cross_checking",
        "adjudicating",
        "generating_report",
        "completed",
    ]

    def __init__(
        self,
        db: Session,
        project_id: str,
        progress_callback: Optional[Callable[[str, int], None]] = None,
        meeting_id: Optional[str] = None,
    ):
        self.db = db
        self.project_id = project_id
        self.meeting_id = meeting_id
        self.progress_callback = progress_callback

    def _q(self, model):
        return scoped_query(self.db, model, self.project_id, self.meeting_id)

    def _missing_documents_for_present(self, present: set[str]) -> list[dict]:
        project = self.db.get(Project, self.project_id)
        pack = get_domain_pack(project)
        if pack.name == "compliance":
            meeting = self.db.get(Meeting, self.meeting_id) if self.meeting_id else None
            meeting_case = dict((meeting.state_json or {}).get("meeting_case") or {}) if meeting else {}
            return check_missing_documents(set(present), domain="compliance", meeting_case=meeting_case)
        return check_missing_documents(set(present))

    def run(self) -> None:
        project = self.db.get(Project, self.project_id)
        if not project:
            raise ValueError("project not found")

        trace = AgentTrace(self.db, self.project_id, self.meeting_id)

        try:
            require_agent_llm()
            files = self._q(FileRecord).all()

            with trace.timed_step("planning"):
                agent_plan = plan_analysis(self.db, self.project_id, files)
            graph = ExecutionGraph.from_plan(agent_plan, files)
            trace.plan(agent_plan, graph.to_dict())

            parsed_docs: list[dict] = []

            if graph.should_run("classifying"):
                self._set_status(project, "classifying")
                trace.step("classifying", "running")
                for i, f in enumerate(files):
                    ext = Path(f.file_name).suffix.lower()
                    classification = classify_document(f.file_name, ext)
                    f.file_type = classification["file_type"]
                    f.document_category = classification["document_category"]
                    f.confidence = classification["confidence"]
                    f.meta_json = classification
                    self.db.commit()
                    if len(files) > 0:
                        self._progress_step_span("classifying", (i + 1) / len(files))
                trace.step("classifying", "completed", {"file_count": len(files)})
            else:
                trace.step("classifying", "skipped", {"reason": "plan_steps"})

            graph = ExecutionGraph.from_plan(agent_plan, files)

            if graph.should_run("parsing"):
                self._set_status(project, "parsing")
                trace.step("parsing", "running")
                for i, f in enumerate(files):
                    content = self._parse_file(f)
                    self._log_pdf_vision_slices(trace, f, content)
                    headers = []
                    if content.get("sheets"):
                        headers = [c["name"] for c in content["sheets"][0].get("columns", [])]
                    reclassify = classify_document(
                        f.file_name, Path(f.file_name).suffix, headers=headers, text=content.get("text_content", "")
                    )
                    if reclassify["confidence"] > (f.confidence or 0):
                        f.document_category = reclassify["document_category"]
                        f.confidence = reclassify["confidence"]
                        f.meta_json = reclassify
                    upsert_parsed_document(
                        self.db,
                        project_id=self.project_id,
                        meeting_id=self.meeting_id,
                        file_id=f.id,
                        document_type=f.document_category,
                        content_json=content,
                        text_content=content.get("text_content", ""),
                    )
                    f.parse_status = "done"
                    self.db.commit()
                    parsed_docs.append(
                        {
                            "file_id": f.id,
                            "file_name": f.file_name,
                            "document_category": f.document_category,
                            "content_json": content,
                            "text_content": content.get("text_content", ""),
                        }
                    )
                    if len(files) > 0:
                        self._progress_step_span("parsing", (i + 1) / len(files))
                parsed_docs = flatten_parsed_docs(parsed_docs)
                trace.step("parsing", "completed", {"parsed": len(parsed_docs)})
            else:
                trace.step("parsing", "skipped", {"reason": "plan_steps"})

            graph = ExecutionGraph.from_plan(agent_plan, files)

            links: list[dict] = []
            entities: list[dict] = []

            if graph.should_run("extracting") and parsed_docs:
                self._set_status(project, "extracting")
                self._progress("extracting")
                trace.step("extracting", "running")
                scoped_delete(self.db, ExtractedEntity, self.project_id, self.meeting_id)
                scoped_delete(self.db, RecordLink, self.project_id, self.meeting_id)
                entities = extract_entities_from_documents(self.project_id, parsed_docs)
                for ent in entities:
                    self.db.add(
                        ExtractedEntity(
                            project_id=ent["project_id"],
                            meeting_id=self.meeting_id,
                            file_id=ent["file_id"],
                            entity_type=ent["entity_type"],
                            entity_value=ent["entity_value"],
                            standard_value=ent.get("standard_value"),
                            source_location=ent.get("source_location"),
                            confidence=ent.get("confidence"),
                        )
                    )
                links = build_record_links(self.project_id, entities)
                for link in links:
                    self.db.add(RecordLink(**{**link, "meeting_id": self.meeting_id}))
                self.db.commit()
                trace.step(
                    "extracting",
                    "completed",
                    {"entities": len(entities), "links": len(links)},
                )
            else:
                trace.step("extracting", "skipped", {"reason": "plan_steps or no parsed docs"})

            all_risks: list[dict] = []
            invoice_rows: list[dict] = []
            expense_rows: list[dict] = []

            if graph.should_run("running_rules") and parsed_docs:
                self._set_status(project, "running_rules")
                self._progress("running_rules")
                trace.step("running_rules", "running", {"focus": sorted(graph.rule_focus_categories)})
                rules = graph.sort_rules(self._load_rules())
                for doc in parsed_docs:
                    if doc["document_category"] in RULE_CATEGORIES:
                        for sheet in doc["content_json"].get("sheets", []):
                            columns_map = {
                                c["name"]: c.get("standard_field")
                                for c in sheet.get("columns", [])
                                if c.get("standard_field")
                            }
                            rows = [
                                {
                                    "row_number": r["row_number"],
                                    "values": r["values"],
                                    "columns_map": columns_map,
                                    "sheet_name": sheet["sheet_name"],
                                }
                                for r in sheet.get("rows", [])
                            ]
                            hits = run_rules_on_rows(
                                rows,
                                rules,
                                {
                                    "file_id": doc["file_id"],
                                    "document_category": doc["document_category"],
                                },
                            )
                            for hit in hits:
                                risk = self._hit_to_risk(hit, doc)
                                all_risks.append(risk)
                            if doc["document_category"] == "expense_detail":
                                expense_rows.extend(rows)
                            if doc["document_category"] in ("expense_detail", "invoice_list"):
                                for r in rows:
                                    vals = r["values"]
                                    inv = vals.get("invoice_number") or vals.get("发票号") or vals.get("发票号码")
                                    if inv:
                                        invoice_rows.append(
                                            {
                                                "invoice_number": str(inv),
                                                "row": r["row_number"],
                                                "file_name": doc["file_name"],
                                            }
                                        )

                    if doc["document_category"] in FIELD_RULE_CATEGORIES:
                        fields = doc["content_json"].get("fields") or {}
                        if fields:
                            hits = run_rules_on_fields(
                                fields,
                                rules,
                                {"file_id": doc["file_id"], "document_category": doc["document_category"]},
                            )
                            for hit in hits:
                                all_risks.append(self._hit_to_risk(hit, doc))

                    if doc["document_category"] == "expense_detail" and graph.should_run_cross("cross_period"):
                        all_risks.extend(
                            detect_cross_period_risks(
                                [
                                    {
                                        **r,
                                        "sheet_name": sheet["sheet_name"],
                                    }
                                    for sheet in doc["content_json"].get("sheets", [])
                                    for r in [
                                        {
                                            "row_number": row["row_number"],
                                            "values": row["values"],
                                        }
                                        for row in sheet.get("rows", [])
                                    ]
                                ],
                                doc["file_id"],
                                doc["file_name"],
                            )
                        )

                compliance_risks = self._run_compliance_domain_checks(project, files, parsed_docs, rules)
                all_risks.extend(compliance_risks)
                trace.step("running_rules", "completed", {"rule_hits": len(all_risks)})
            else:
                trace.step("running_rules", "skipped", {"reason": "plan_steps or no parsed docs"})

            if graph.should_run("cross_checking") and parsed_docs:
                self._set_status(project, "cross_checking")
                self._progress("cross_checking")
                trace.step("cross_checking", "running", {"modules": sorted(graph.cross_modules)})
                if graph.should_run_cross("amounts"):
                    all_risks.extend(cross_check_amounts(parsed_docs))
                if graph.should_run_cross("duplicates"):
                    all_risks.extend(detect_duplicate_invoices(invoice_rows))
                if graph.should_run_cross("record_links"):
                    all_risks.extend(links_to_cross_risks(links, parsed_docs))
                if graph.should_run_cross("three_way"):
                    all_risks.extend(detect_three_way_gaps(parsed_docs, links))
                if graph.should_run_cross("anomalies"):
                    all_risks.extend(detect_amount_anomalies(expense_rows))
                trace.step("cross_checking", "completed", {"total_risks": len(all_risks)})
            else:
                trace.step("cross_checking", "skipped", {"reason": "plan_steps or no parsed docs"})

            present = {d["document_category"] for d in parsed_docs}
            missing = self._missing_documents_for_present(present)

            if graph.should_run("adjudicating"):
                self._set_status(project, "adjudicating")
                self._progress("adjudicating")
                trace.step("adjudicating", "running", {"risk_count": len(all_risks)})
                all_risks = adjudicate_risks(self.db, all_risks, agent_plan)
                for r in all_risks:
                    if r.get("risk_level"):
                        scores = calculate_risk_score(
                            r["risk_level"],
                            r.get("evidence_json") or {},
                            r.get("confidence", 0.9),
                        )
                        r["risk_score"] = scores["total_score"]
                        r["risk_level"] = scores["risk_level"]
                trace.step(
                    "adjudicating",
                    "completed",
                    {"agent_mode": agent_plan.get("agent_mode"), "risk_count": len(all_risks)},
                )
            else:
                trace.step("adjudicating", "skipped", {"reason": "plan_steps"})

            scoped_delete(self.db, Risk, self.project_id, self.meeting_id)
            for r in all_risks:
                risk_obj = Risk(
                    project_id=self.project_id,
                    meeting_id=self.meeting_id,
                    risk_id=r.get("risk_id") or f"RISK-{uuid.uuid4().hex[:8]}",
                    risk_category=r["risk_category"],
                    risk_subcategory=r.get("risk_subcategory"),
                    risk_level=r["risk_level"],
                    risk_score=r.get("risk_score", 0),
                    source_file_id=r.get("source_file_id"),
                    source_location_json=r.get("source_location_json"),
                    related_files=r.get("related_files"),
                    problem=r["problem"],
                    evidence_json=r["evidence_json"],
                    rule_triggered=r.get("rule_triggered"),
                    analysis=r.get("analysis"),
                    suggestion=r["suggestion"],
                    correction_action=r.get("correction_action", "待处理"),
                    manual_review_required=r.get("manual_review_required", False),
                    confidence=r.get("confidence", 0.9),
                    status="pending",
                )
                self.db.add(risk_obj)
            self.db.commit()

            if graph.should_run("generating_report"):
                self._set_status(project, "generating_report")
                self._progress("generating_report")
                self._generate_outputs(project, missing)
            project.summary = self._build_summary(all_risks, missing)
            prior = dict(project.state_json or {})
            project.state_json = {
                **prior,
                "agent_plan": agent_plan,
                "execution_graph": graph.to_dict(),
                "missing_documents": missing,
                "risk_count": len(all_risks),
                "present_categories": list(present),
                "entity_count": len(entities),
                "link_count": len(links),
                "processed_file_ids": [f.id for f in files],
                "execution_mode": "pipeline",
            }
            self._set_status(project, "completed")
            self._progress("completed")
            trace.step("workflow", "completed", {"risk_count": len(all_risks)})
        except Exception as exc:
            self._set_status(project, "failed")
            trace.step("workflow", "failed", {"error": str(exc)})
            raise

    def run_react(self) -> None:
        """ReAct 外环：LLM 逐步调度内环流水线工具。"""
        from app.services.agent.pipeline_executor import PipelineExecutor

        trace = AgentTrace(self.db, self.project_id, self.meeting_id)
        try:
            require_agent_llm()
            executor = PipelineExecutor(
                self.db, self.project_id, self.progress_callback, trace, self.meeting_id
            )
            executor.run_react()
        except Exception as exc:
            project = self.db.get(Project, self.project_id)
            if project:
                self._set_status(project, "failed")
            trace.step("workflow", "failed", {"error": str(exc), "mode": "react"})
            raise

    def run_orchestrator(self) -> None:
        """Orchestrator 外环：主 Agent 拆解任务并委派子 Agent。"""
        from app.services.agent.orchestrator import MissionOrchestrator

        trace = AgentTrace(self.db, self.project_id, self.meeting_id)
        try:
            MissionOrchestrator(
                self.db, self.project_id, self.progress_callback, trace, self.meeting_id
            ).run()
        except Exception as exc:
            project = self.db.get(Project, self.project_id)
            if project:
                self._set_status(project, "failed")
            trace.step("workflow", "failed", {"error": str(exc), "mode": "orchestrator"})
            raise

    def run_partial(self, scope: str) -> None:
        """局部重跑：cross_checking（交叉比对+研判+报告）或 adjudicating（仅研判+报告）。"""
        if scope not in ("cross_checking", "adjudicating"):
            raise ValueError(f"unsupported partial scope: {scope}")

        project = self.db.get(Project, self.project_id)
        if not project:
            raise ValueError("project not found")

        state = project.state_json or {}
        agent_plan = state.get("agent_plan") or {}
        if not agent_plan:
            raise ValueError("缺少 agent_plan，请先完成全量分析")

        trace = AgentTrace(self.db, self.project_id, self.meeting_id)
        trace.step("runtime", "running", {"scope": scope, "kind": "partial_rerun"})

        try:
            require_agent_llm()
            files = self._q(FileRecord).all()
            graph = ExecutionGraph.from_plan(agent_plan, files)
            parsed_docs = self._load_parsed_docs_from_db()
            links = self._load_links_from_db()
            missing = state.get("missing_documents") or []

            if scope == "adjudicating":
                all_risks = self._risks_from_db()
                if not all_risks:
                    raise ValueError("无风险可研判")
            else:
                base_risks = [
                    self._risk_orm_to_dict(r)
                    for r in self._q(Risk).all()
                    if not _is_cross_derived_risk(r.rule_triggered)
                ]
                invoice_rows, expense_rows = self._collect_invoice_expense_rows(parsed_docs)
                cross_risks: list[dict] = []
                self._set_status(project, "cross_checking")
                self._progress("cross_checking")
                if graph.should_run_cross("amounts"):
                    cross_risks.extend(cross_check_amounts(parsed_docs))
                if graph.should_run_cross("duplicates"):
                    cross_risks.extend(detect_duplicate_invoices(invoice_rows))
                if graph.should_run_cross("record_links"):
                    cross_risks.extend(links_to_cross_risks(links, parsed_docs))
                if graph.should_run_cross("three_way"):
                    cross_risks.extend(detect_three_way_gaps(parsed_docs, links))
                if graph.should_run_cross("anomalies"):
                    cross_risks.extend(detect_amount_anomalies(expense_rows))
                all_risks = base_risks + cross_risks

            self._set_status(project, "adjudicating")
            self._progress("adjudicating")
            all_risks = adjudicate_risks(self.db, all_risks, agent_plan)
            for r in all_risks:
                if r.get("risk_level"):
                    scores = calculate_risk_score(
                        r["risk_level"],
                        r.get("evidence_json") or {},
                        r.get("confidence", 0.9),
                    )
                    r["risk_score"] = scores["total_score"]
                    r["risk_level"] = scores["risk_level"]

            self._persist_risks(all_risks)
            self._set_status(project, "generating_report")
            self._progress("generating_report")
            self._generate_outputs(project, missing)
            project.summary = self._build_summary(all_risks, missing)
            state["partial_rerun"] = {"scope": scope, "risk_count": len(all_risks)}
            project.state_json = state
            self._set_status(project, "completed")
            self._progress("completed")
            trace.step("runtime", "completed", {"scope": scope, "risk_count": len(all_risks)})
        except Exception as exc:
            self._set_status(project, "failed")
            trace.step("runtime", "failed", {"scope": scope, "error": str(exc)})
            raise

    def run_incremental(self) -> None:
        """补资料后增量分析：仅解析新文件，重跑受影响阶段。"""
        project = self.db.get(Project, self.project_id)
        if not project:
            raise ValueError("project not found")

        state = dict(project.state_json or {})
        base_plan = state.get("agent_plan") or {}
        if not base_plan:
            raise ValueError("缺少历史 agent_plan，请先完成全量分析")

        files = self._q(FileRecord).all()
        diff = diff_uploaded_files(state, files)
        if not diff.new_file_ids:
            raise ValueError("没有检测到新增资料")

        agent_plan = enrich_plan_with_sub_agents(
            build_incremental_plan(base_plan, diff),
            files,
        )
        graph = ExecutionGraph.from_plan(agent_plan, files)
        trace = AgentTrace(self.db, self.project_id, self.meeting_id)
        trace.plan(agent_plan, graph.to_dict())
        trace.step("runtime", "running", {"scope": "incremental", "new_files": len(diff.new_file_ids)})

        try:
            require_agent_llm()
            new_files = [f for f in files if f.id in diff.new_file_ids]

            if graph.should_run("classifying"):
                self._set_status(project, "classifying")
                for i, f in enumerate(new_files):
                    ext = Path(f.file_name).suffix.lower()
                    classification = classify_document(f.file_name, ext)
                    f.file_type = classification["file_type"]
                    f.document_category = classification["document_category"]
                    f.confidence = classification["confidence"]
                    f.meta_json = classification
                    self.db.commit()
                    if len(new_files) > 0:
                        self._progress_step_span("classifying", (i + 1) / len(new_files))

            parsed_docs = self._load_parsed_docs_from_db()

            if graph.should_run("parsing"):
                self._set_status(project, "parsing")
                for i, f in enumerate(new_files):
                    content = self._parse_file(f)
                    self._log_pdf_vision_slices(trace, f, content)
                    headers = []
                    if content.get("sheets"):
                        headers = [c["name"] for c in content["sheets"][0].get("columns", [])]
                    reclassify = classify_document(
                        f.file_name, Path(f.file_name).suffix, headers=headers, text=content.get("text_content", "")
                    )
                    if reclassify["confidence"] > (f.confidence or 0):
                        f.document_category = reclassify["document_category"]
                        f.confidence = reclassify["confidence"]
                        f.meta_json = reclassify
                    upsert_parsed_document(
                        self.db,
                        project_id=self.project_id,
                        meeting_id=self.meeting_id,
                        file_id=f.id,
                        document_type=f.document_category,
                        content_json=content,
                        text_content=content.get("text_content", ""),
                    )
                    f.parse_status = "done"
                    self.db.commit()
                    if len(new_files) > 0:
                        self._progress_step_span("parsing", (i + 1) / len(new_files))
                parsed_docs = self._load_parsed_docs_from_db()

            graph = ExecutionGraph.from_plan(agent_plan, files)
            links: list[dict] = []
            entities: list[dict] = []
            all_risks: list[dict] = []
            invoice_rows: list[dict] = []
            expense_rows: list[dict] = []

            if graph.should_run("extracting") and parsed_docs:
                self._set_status(project, "extracting")
                self._progress("extracting")
                scoped_delete(self.db, ExtractedEntity, self.project_id, self.meeting_id)
                scoped_delete(self.db, RecordLink, self.project_id, self.meeting_id)
                entities = extract_entities_from_documents(self.project_id, parsed_docs)
                for ent in entities:
                    self.db.add(
                        ExtractedEntity(
                            project_id=ent["project_id"],
                            meeting_id=self.meeting_id,
                            file_id=ent["file_id"],
                            entity_type=ent["entity_type"],
                            entity_value=ent["entity_value"],
                            standard_value=ent.get("standard_value"),
                            source_location=ent.get("source_location"),
                            confidence=ent.get("confidence"),
                        )
                    )
                links = build_record_links(self.project_id, entities)
                for link in links:
                    self.db.add(RecordLink(**{**link, "meeting_id": self.meeting_id}))
                self.db.commit()

            if graph.should_run("running_rules") and parsed_docs:
                self._set_status(project, "running_rules")
                self._progress("running_rules")
                rules = graph.sort_rules(self._load_rules())
                for doc in parsed_docs:
                    if doc["document_category"] in RULE_CATEGORIES:
                        for sheet in doc["content_json"].get("sheets", []):
                            columns_map = {
                                c["name"]: c.get("standard_field")
                                for c in sheet.get("columns", [])
                                if c.get("standard_field")
                            }
                            rows = [
                                {
                                    "row_number": r["row_number"],
                                    "values": r["values"],
                                    "columns_map": columns_map,
                                    "sheet_name": sheet["sheet_name"],
                                }
                                for r in sheet.get("rows", [])
                            ]
                            hits = run_rules_on_rows(
                                rows,
                                rules,
                                {"file_id": doc["file_id"], "document_category": doc["document_category"]},
                            )
                            for hit in hits:
                                all_risks.append(self._hit_to_risk(hit, doc))
                            if doc["document_category"] == "expense_detail":
                                expense_rows.extend(rows)
                            if doc["document_category"] in ("expense_detail", "invoice_list"):
                                for r in rows:
                                    vals = r["values"]
                                    inv = vals.get("invoice_number") or vals.get("发票号") or vals.get("发票号码")
                                    if inv:
                                        invoice_rows.append(
                                            {
                                                "invoice_number": str(inv),
                                                "row": r["row_number"],
                                                "file_name": doc["file_name"],
                                            }
                                        )
                    if doc["document_category"] in FIELD_RULE_CATEGORIES:
                        fields = doc["content_json"].get("fields") or {}
                        if fields:
                            hits = run_rules_on_fields(
                                fields,
                                rules,
                                {"file_id": doc["file_id"], "document_category": doc["document_category"]},
                            )
                            for hit in hits:
                                all_risks.append(self._hit_to_risk(hit, doc))

            if graph.should_run("cross_checking") and parsed_docs:
                self._set_status(project, "cross_checking")
                self._progress("cross_checking")
                if graph.should_run_cross("amounts"):
                    all_risks.extend(cross_check_amounts(parsed_docs))
                if graph.should_run_cross("duplicates"):
                    all_risks.extend(detect_duplicate_invoices(invoice_rows))
                if graph.should_run_cross("record_links"):
                    all_risks.extend(links_to_cross_risks(links, parsed_docs))
                if graph.should_run_cross("three_way"):
                    all_risks.extend(detect_three_way_gaps(parsed_docs, links))
                if graph.should_run_cross("anomalies"):
                    all_risks.extend(detect_amount_anomalies(expense_rows))

            present = {d["document_category"] for d in parsed_docs}
            missing = self._missing_documents_for_present(present)

            if graph.should_run("adjudicating"):
                self._set_status(project, "adjudicating")
                self._progress("adjudicating")
                all_risks = adjudicate_risks(self.db, all_risks, agent_plan)
                for r in all_risks:
                    if r.get("risk_level"):
                        scores = calculate_risk_score(
                            r["risk_level"],
                            r.get("evidence_json") or {},
                            r.get("confidence", 0.9),
                        )
                        r["risk_score"] = scores["total_score"]
                        r["risk_level"] = scores["risk_level"]

            self._persist_risks(all_risks)

            if graph.should_run("generating_report"):
                self._set_status(project, "generating_report")
                self._progress("generating_report")
                self._generate_outputs(project, missing)

            project.summary = self._build_summary(all_risks, missing)
            state.update(
                {
                    "agent_plan": agent_plan,
                    "execution_graph": graph.to_dict(),
                    "missing_documents": missing,
                    "risk_count": len(all_risks),
                    "present_categories": list(present),
                    "entity_count": len(entities),
                    "link_count": len(links),
                    "processed_file_ids": list(diff.current_file_ids),
                    "last_incremental": {
                        "new_file_ids": diff.new_file_ids,
                        "new_categories": sorted(diff.new_categories),
                    },
                }
            )
            project.state_json = state
            self._set_status(project, "completed")
            self._progress("completed")
            trace.step(
                "runtime",
                "completed",
                {"scope": "incremental", "new_files": len(diff.new_file_ids), "risk_count": len(all_risks)},
            )
        except Exception as exc:
            self._set_status(project, "failed")
            trace.step("runtime", "failed", {"scope": "incremental", "error": str(exc)})
            raise

    def regenerate_outputs_only(self) -> None:
        """人工复核后仅重新生成交付物，不重跑分析。"""
        project = self.db.get(Project, self.project_id)
        if not project:
            raise ValueError("project not found")
        missing = (project.state_json or {}).get("missing_documents", [])
        self._set_status(project, "generating_report")
        self._generate_outputs(project, missing)
        self._set_status(project, "completed")
        self._log("regenerate_outputs", "completed", {})

    def _load_parsed_docs_from_db(self) -> list[dict]:
        docs = []
        q = self._q(ParsedDocument)
        for pd in q.all():
            f = self.db.get(FileRecord, pd.file_id)
            if not f:
                continue
            docs.append(
                {
                    "file_id": f.id,
                    "file_name": f.file_name,
                    "document_category": f.document_category,
                    "content_json": pd.content_json,
                    "text_content": pd.text_content or "",
                }
            )
        return flatten_parsed_docs(docs)

    def _load_links_from_db(self) -> list[dict]:
        return [
            {
                "project_id": l.project_id,
                "link_type": l.link_type,
                "source_file_id": l.source_file_id,
                "target_file_id": l.target_file_id,
                "link_keys": l.link_keys,
                "confidence": l.confidence,
                "match_method": l.match_method,
            }
            for l in self._q(RecordLink).all()
        ]

    def _risk_orm_to_dict(self, r: Risk) -> dict:
        return {
            "risk_id": r.risk_id,
            "risk_category": r.risk_category,
            "risk_subcategory": r.risk_subcategory,
            "risk_level": r.risk_level,
            "risk_score": r.risk_score,
            "source_file_id": r.source_file_id,
            "source_location_json": r.source_location_json,
            "related_files": r.related_files,
            "problem": r.problem,
            "evidence_json": r.evidence_json,
            "rule_triggered": r.rule_triggered,
            "analysis": r.analysis,
            "suggestion": r.suggestion,
            "correction_action": r.correction_action,
            "manual_review_required": r.manual_review_required,
            "confidence": r.confidence,
        }

    def _risks_from_db(self) -> list[dict]:
        return [self._risk_orm_to_dict(r) for r in self._q(Risk).all()]

    def _collect_invoice_expense_rows(self, parsed_docs: list[dict]) -> tuple[list[dict], list[dict]]:
        invoice_rows: list[dict] = []
        expense_rows: list[dict] = []
        for doc in parsed_docs:
            if doc["document_category"] not in RULE_CATEGORIES:
                continue
            for sheet in doc["content_json"].get("sheets", []):
                rows = [
                    {
                        "row_number": r["row_number"],
                        "values": r["values"],
                        "sheet_name": sheet["sheet_name"],
                    }
                    for r in sheet.get("rows", [])
                ]
                if doc["document_category"] == "expense_detail":
                    expense_rows.extend(rows)
                if doc["document_category"] in ("expense_detail", "invoice_list"):
                    for r in rows:
                        vals = r["values"]
                        inv = vals.get("invoice_number") or vals.get("发票号") or vals.get("发票号码")
                        if inv:
                            invoice_rows.append(
                                {
                                    "invoice_number": str(inv),
                                    "row": r["row_number"],
                                    "file_name": doc["file_name"],
                                }
                            )
        return invoice_rows, expense_rows

    def _run_compliance_domain_checks(
        self,
        project: Project,
        files: list[FileRecord],
        parsed_docs: list[dict],
        rules: list[Rule],
    ) -> list[dict]:
        if get_domain_pack(project).name != "compliance":
            return []

        meeting = self.db.get(Meeting, self.meeting_id) if self.meeting_id else None
        state_owner = meeting or project
        state = dict(state_owner.state_json or {})
        meeting_case = dict(state.get("meeting_case") or {})
        facts = build_compliance_case_facts(meeting_case, files, parsed_docs)
        hits = run_compliance_checks(facts, rules)
        observation_type = str(facts.get("observation_type") or meeting_case.get("observation_type") or "远程观察")
        risks = generate_finding_narratives(hits, {**meeting_case, **facts}, observation_type)

        state["meeting_case"] = {**meeting_case, **facts, "finding_count": len(risks)}
        state["agent_domain"] = "compliance"
        state_owner.state_json = state
        if meeting:
            meeting.observation_type = facts.get("observation_type") or meeting.observation_type
            meeting.meeting_type = facts.get("meeting_type") or meeting.meeting_type
            meeting.meeting_code = facts.get("meeting_code") or meeting.meeting_code
        self.db.commit()
        return risks

    def _persist_risks(self, all_risks: list[dict]) -> None:
        scoped_delete(self.db, Risk, self.project_id, self.meeting_id)
        for r in all_risks:
            self.db.add(
                Risk(
                    project_id=self.project_id,
                    meeting_id=self.meeting_id,
                    risk_id=r.get("risk_id") or f"RISK-{uuid.uuid4().hex[:8]}",
                    risk_category=r["risk_category"],
                    risk_subcategory=r.get("risk_subcategory"),
                    risk_level=r["risk_level"],
                    risk_score=r.get("risk_score", 0),
                    source_file_id=r.get("source_file_id"),
                    source_location_json=r.get("source_location_json"),
                    related_files=r.get("related_files"),
                    problem=r["problem"],
                    evidence_json=r["evidence_json"],
                    rule_triggered=r.get("rule_triggered"),
                    analysis=r.get("analysis"),
                    suggestion=r["suggestion"],
                    correction_action=r.get("correction_action", "待处理"),
                    manual_review_required=r.get("manual_review_required", False),
                    confidence=r.get("confidence", 0.9),
                    status="pending",
                )
            )
        self.db.commit()

    def _parse_file(self, f: FileRecord) -> dict:
        path = Path(f.storage_path)
        if f.file_type == "excel" or path.suffix.lower() in (".xlsx", ".xls", ".csv"):
            return parse_excel(path)
        if f.file_type == "pdf" or path.suffix.lower() == ".pdf":
            project = self.db.get(Project, self.project_id)
            pack = get_domain_pack(project)
            result = ingest_pdf_hybrid(
                path,
                f.document_category or "unknown",
                f.file_name,
                domain=pack.name,
            )
            content = result.to_content_json()
            if "contract" in f.document_category or "合同" in f.file_name:
                content["fields"] = extract_contract_fields(content)
                f.document_category = "contract"
            return content
        if f.file_type == "word" or path.suffix.lower() in (".docx", ".doc"):
            content = parse_word(path)
            f.document_category = "contract"
            return content
        if f.file_type == "image" or path.suffix.lower() in (".jpg", ".jpeg", ".png"):
            project = self.db.get(Project, self.project_id)
            if get_domain_pack(project).name == "compliance":
                from app.services.domain.compliance.compliance_vision import analyze_compliance_image

                return analyze_compliance_image(path, f.document_category or "unknown", f.file_name)
            return parse_image(path)
        return {"file_type": "unknown", "text_content": ""}

    def _log_pdf_vision_slices(self, trace: AgentTrace, f: FileRecord, content: dict) -> None:
        """Expose PDF image-page OCR as vision-agent work, even though PDF ingest starts in parsing."""
        if content.get("file_type") != "pdf":
            return
        slices = content.get("vision_slices") or []
        if not slices:
            return
        page_count = len({s.get("page_number") for s in slices if s.get("page_number")})
        modes = sorted({str(s.get("ingest_mode") or "") for s in slices if s.get("ingest_mode")})
        trace.log(
            "vision_agent",
            "completed",
            kind="vision_agent",
            name="视觉 Agent",
            message=f"PDF 图像页识别 {f.file_name}",
            detail={
                "file_id": f.id,
                "file_name": f.file_name,
                "category": f.document_category,
                "pdf_type": content.get("pdf_type"),
                "ingest_mode": content.get("ingest_mode"),
                "ocr_engine": content.get("ocr_engine"),
                "vision_slice_count": len(slices),
                "vision_page_count": page_count,
                "vision_modes": modes,
            },
        )

    def _load_rules(self) -> list[dict]:
        db_rules = (
            self.db.query(Rule)
            .filter_by(enabled=True)
            .order_by(Rule.priority.desc(), Rule.rule_id)
            .all()
        )
        return [
            {
                "rule_id": r.rule_id,
                "rule_name": r.rule_name,
                "risk_category": r.risk_category,
                "risk_level": r.risk_level,
                "applicable_document_type": r.applicable_document_type,
                "condition_json": r.condition_json,
                "evidence_fields": r.evidence_fields or [],
                "suggestion_template": r.suggestion_template,
                "manual_review_required": r.manual_review_required,
                "enabled": r.enabled,
            }
            for r in db_rules
        ]

    def _hit_to_risk(self, hit: dict, doc: dict) -> dict:
        rule = hit["rule"]
        evidence = hit["evidence"]
        row_ctx = hit["row_ctx"]
        confidence = 0.9
        if doc.get("document_category") == "unknown":
            confidence = 0.6
        scores = calculate_risk_score(rule["risk_level"], evidence, confidence)
        loc_col = next((k for k in evidence if k), None)
        return {
            "risk_id": f"{rule['rule_id']}-{row_ctx.get('row_number')}",
            "risk_category": rule["risk_category"],
            "risk_subcategory": rule["rule_name"],
            "risk_level": scores["risk_level"],
            "risk_score": scores["total_score"],
            "source_file_id": doc["file_id"],
            "source_location_json": {
                "sheet": row_ctx.get("sheet_name"),
                "row": row_ctx.get("row_number"),
                "column": loc_col,
            },
            "related_files": [doc["file_name"]],
            "problem": rule["rule_name"],
            "evidence_json": evidence,
            "rule_triggered": rule["rule_id"],
            "analysis": None,
            "suggestion": rule["suggestion_template"],
            "correction_action": "标记为需补充资料" if rule["manual_review_required"] else "待核实",
            "manual_review_required": rule["manual_review_required"] or confidence < 0.75,
            "confidence": confidence,
        }

    def _generate_outputs(self, project: Project, missing: list[dict]) -> None:
        rq = self._q(Risk).filter(Risk.status != "dismissed")
        if self.meeting_id:
            rq = rq.filter_by(meeting_id=self.meeting_id)
        risks = rq.all()
        risk_dicts = [
            {
                "risk_id": r.risk_id,
                "risk_level": r.risk_level,
                "risk_score": r.risk_score,
                "risk_category": r.risk_category,
                "risk_subcategory": r.risk_subcategory,
                "problem": r.problem,
                "suggestion": r.suggestion,
                "evidence_json": r.evidence_json,
                "rule_triggered": r.rule_triggered,
                "manual_review_required": r.manual_review_required,
                "status": r.status,
                "source_file_id": r.source_file_id,
                "source_location_json": r.source_location_json,
                "correction_action": r.correction_action,
            }
            for r in risks
        ]

        out_dir = settings.storage_path / "outputs" / self.project_id
        if self.meeting_id:
            out_dir = out_dir / self.meeting_id
        out_dir.mkdir(parents=True, exist_ok=True)

        from app.models import Meeting

        meeting = self.db.get(Meeting, self.meeting_id) if self.meeting_id else None
        state = dict((meeting.state_json if meeting else project.state_json) or {})
        domain = state.get("agent_domain") or settings.agent_domain
        is_compliance = domain == "compliance"
        files_q = self._q(FileRecord)
        if self.meeting_id:
            files_q = files_q.filter_by(meeting_id=self.meeting_id)
        files = files_q.all()
        file_names = {f.id: f.file_name for f in files}
        present = state.get("present_categories") or {
            f.document_category for f in files if f.document_category and f.document_category != "unknown"
        }
        pack = get_domain_pack(project)
        if is_compliance and pack.name == "compliance":
            meeting_case_for_missing = state.get("meeting_case") or {}
            missing = check_missing_documents(
                set(present),
                domain="compliance",
                meeting_case=meeting_case_for_missing,
            )

        if is_compliance:
            meeting_case = state.get("meeting_case") or {}
            parsed_materials = collect_parsed_materials(self.db, self.project_id, self.meeting_id)
            bundle = build_compliance_deliverable_bundle(
                out_dir,
                project.name,
                risk_dicts,
                missing,
                meeting_case=meeting_case,
                runtime=state.get("runtime"),
                file_names=file_names,
                parsed_materials=parsed_materials,
            )
            fixed_template_path = bundle.get("fixed_template_excel")
            fixed_template_patch = (
                _meeting_case_patch_from_fixed_template(fixed_template_path)
                if fixed_template_path
                else {}
            )
            template_quality = {}
            quality_json_path = bundle.get("fixed_template_quality_json")
            if quality_json_path and Path(quality_json_path).exists():
                template_quality = compact_template_quality(load_template_quality_json(Path(quality_json_path)))
            if fixed_template_patch:
                state = dict(state)
                next_meeting_case = dict(state.get("meeting_case") or {})
                next_meeting_case.update(fixed_template_patch)
                state["meeting_case"] = next_meeting_case
                if meeting:
                    meeting.state_json = state
            if template_quality:
                deliverable = dict((meeting.deliverable_json if meeting else None) or state.get("deliverable") or {})
                deliverable.setdefault("status", "pending")
                deliverable.setdefault("comment", "")
                deliverable["template_quality"] = template_quality
                state = dict(state)
                state["deliverable"] = deliverable
                if meeting:
                    meeting.deliverable_json = deliverable
                    meeting.state_json = state
            output_q = self.db.query(Output).filter_by(project_id=project.id)
            if self.meeting_id:
                output_q = output_q.filter_by(meeting_id=self.meeting_id)
            output_q.delete(synchronize_session=False)
            self.db.commit()
            for otype in PRIMARY_DELIVERABLE_TYPES:
                path = bundle.get(otype)
                if path:
                    self._save_output(project, otype, path)
        else:
            excel_path = out_dir / "风险清单.xlsx"
            generate_risk_excel(risk_dicts, excel_path)
            self._save_output(project, "risk_excel", excel_path)
            pdf_path = out_dir / "风险评估报告.pdf"
            generate_pdf_report(project.name, risk_dicts, missing, pdf_path)
            self._save_output(project, "risk_pdf", pdf_path)
            missing_path = out_dir / "缺件清单.xlsx"
            generate_generic_missing_excel(missing, missing_path)
            self._save_output(project, "missing_docs", missing_path)
            correction_path = out_dir / "整改建议清单.xlsx"
            generate_generic_correction_excel(risk_dicts, correction_path)
            self._save_output(project, "correction_list", correction_path)
        if not is_compliance:
            for f in files:
                if f.file_type != "excel":
                    continue
                if Path(f.file_name).suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
                    continue
                annotated = out_dir / f"批注_{f.file_name}"
                file_risks = [rd for rd in risk_dicts if rd.get("source_location_json")]
                source_path = _existing_file_path(f.storage_path)
                if not source_path:
                    continue
                annotate_excel(source_path, file_risks, annotated)
                self._save_output(project, "annotated_excel", annotated)

            for f in files:
                if f.file_type != "image":
                    continue
                ann_path = out_dir / f"标注_{f.file_name}"
                anns = [{"bbox": None, "text": r.problem} for r in risks if r.source_file_id == f.id]
                if anns:
                    source_path = _existing_file_path(f.storage_path)
                    if not source_path:
                        continue
                    annotate_image(source_path, anns, ann_path)
                    self._save_output(project, "annotated_image", ann_path)

        st = dict(project.state_json or {})
        st["missing_documents"] = missing
        if is_compliance:
            st["present_categories"] = sorted(present)
            if "meeting_case" in state:
                st["meeting_case"] = state["meeting_case"]
        project.state_json = st
        project.summary = self._build_summary(risk_dicts, missing)
        self.db.commit()

        # 清理历史 JSON 交付物（已不再对外提供）
        for stale in out_dir.glob("*.json"):
            if stale.name in ("补充资料清单.json", "更正建议清单.json"):
                stale.unlink(missing_ok=True)
        oq = self._q(Output).filter(Output.file_name.like("%.json"))
        oq.delete(synchronize_session=False)
        self.db.commit()

    def _save_output(self, project: Project, output_type: str, path: Path) -> None:
        dq = self.db.query(Output).filter_by(project_id=project.id, output_type=output_type)
        if self.meeting_id:
            dq = dq.filter_by(meeting_id=self.meeting_id)
        dq.delete(synchronize_session=False)
        self.db.add(
            Output(
                project_id=project.id,
                meeting_id=self.meeting_id,
                output_type=output_type,
                file_name=path.name,
                storage_path=str(path),
            )
        )
        self.db.commit()

    def _build_summary(self, risks: list[dict], missing: list[dict]) -> str:
        high = sum(1 for r in risks if r.get("risk_level") == "高")
        mid = sum(1 for r in risks if r.get("risk_level") == "中")
        low = sum(1 for r in risks if r.get("risk_level") == "低")
        return f"共识别风险 {len(risks)} 项（高 {high} / 中 {mid} / 低 {low}），缺失资料 {len(missing)} 类。"

    def _set_status(self, project: Project, status: str) -> None:
        project.status = status
        self.db.commit()
        if status in STEP_PROGRESS:
            self._progress(status)

    def _progress(self, step: str) -> None:
        pct = STEP_PROGRESS.get(step, 0)
        if self.progress_callback:
            self.progress_callback(step, pct)

    def _progress_step_span(self, step: str, fraction: float) -> None:
        """步骤内细粒度进度（如逐文件解析）。"""
        if not self.progress_callback:
            return
        base = STEP_PROGRESS.get(step, 0)
        step_keys = [s for s in self.STEPS if s in STEP_PROGRESS]
        try:
            idx = step_keys.index(step)
            nxt = STEP_PROGRESS[step_keys[idx + 1]] if idx + 1 < len(step_keys) else 100
        except ValueError:
            nxt = min(100, base + 15)
        pct = int(base + (nxt - base) * max(0.0, min(1.0, fraction)))
        self.progress_callback(step, pct)

    def _log(self, step: str, status: str, detail: dict | None = None) -> None:
        started = time.time()
        detail_json = dict(detail or {})
        detail_json.setdefault("code_location", trace_code_location())
        self.db.add(
            AgentRunLog(
                project_id=self.project_id,
                meeting_id=self.meeting_id,
                step=step,
                status=status,
                detail_json=detail_json,
                duration_ms=int((time.time() - started) * 1000),
            )
        )
        self.db.commit()


async def enrich_risks_with_llm(db: Session, project_id: str) -> None:
    """兼容旧调用：研判已在 workflow.adjudicating 步骤完成。"""
    return

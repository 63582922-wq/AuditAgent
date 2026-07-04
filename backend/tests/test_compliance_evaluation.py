from __future__ import annotations

from openpyxl import Workbook

from app.models import FileRecord, Meeting, Output, Project
from app.services.evaluation.compliance_eval import (
    build_db_compliance_snapshot,
    evaluate_compliance_snapshot,
    load_compliance_eval_cases,
)


def test_fx_eval_baselines_define_core_case_expectations() -> None:
    cases = load_compliance_eval_cases()

    assert {case.case_id for case in cases} == {"A1P260307357", "SMS202606090070"}

    a1p = next(case for case in cases if case.case_id == "A1P260307357")
    sms = next(case for case in cases if case.case_id == "SMS202606090070")

    assert a1p.expected_facts["actual_sign_in_count"] == 7
    assert a1p.expected_facts["material_code"] == "M-CN-00013658"
    assert sms.expected_facts["attendance_source"] == "watch_record"
    assert sms.expected_facts["actual_sign_in_count"] is None
    assert sms.expected_facts["watch_record_count"] == 60
    assert sms.expected_facts["max_attendee_count"] == 15
    assert sms.expected_facts["zoom_peak_count"] == 15
    assert sms.expected_facts["total_attendance_expression"] == "5+60人次"


def test_compliance_eval_snapshot_passes_when_actual_matches_baseline() -> None:
    case = next(case for case in load_compliance_eval_cases() if case.case_id == "A1P260307357")
    report = evaluate_compliance_snapshot(
        case,
        {
            "file_count": 28,
            "file_categories": {
                "meeting_screenshot": 12,
                "coordination_sms": 5,
                "sign_in_record": 3,
                "a1_meeting_export": 1,
                "observation_confirmation": 3,
                "presentation_material": 2,
                "meeting_agenda": 1,
                "speaker_profile": 1,
            },
            "facts": {
                "actual_sign_in_count": 7,
                "planned_attendees": 6,
                "attendance_delta": 1,
                "material_code": "M-CN-00013658",
                "presentation_topic": "SALWEEN研究：重塑PCV初治患者一线新标准",
            },
            "template": {"columns": 143, "filled_row2": 143},
            "outputs": ["fixed_template_excel", "deliverable_package"],
        },
    )

    assert report["passed"] is True
    assert report["critical_failures"] == 0
    assert {item["check_id"] for item in report["checks"] if item["passed"]} >= {
        "file_count",
        "category:sign_in_record",
        "fact:actual_sign_in_count",
        "template:columns",
        "output:fixed_template_excel",
    }


def test_compliance_eval_snapshot_flags_wrong_sign_in_and_delivery_scope() -> None:
    case = next(case for case in load_compliance_eval_cases() if case.case_id == "SMS202606090070")
    report = evaluate_compliance_snapshot(
        case,
        {
            "file_count": 56,
            "file_categories": {"sign_in_record": 1, "meeting_screenshot": 44},
            "facts": {
                "attendance_source": "watch_record",
                "watch_record_count": 60,
                "actual_sign_in_count": 0,
            },
            "template": {"columns": 143, "filled_row2": 120},
            "outputs": ["fixed_template_excel", "finding_excel", "deliverable_package"],
        },
    )

    failed = {item["check_id"]: item for item in report["checks"] if not item["passed"]}
    assert report["passed"] is False
    assert failed["fact:actual_sign_in_count"]["severity"] == "critical"
    assert failed["template:filled_row2"]["severity"] == "critical"
    assert failed["output_scope:no_loose_outputs"]["severity"] == "warning"


def test_compliance_eval_diagnoses_failed_checks_to_files_and_fields() -> None:
    case = next(case for case in load_compliance_eval_cases() if case.case_id == "A1P260307357")
    report = evaluate_compliance_snapshot(
        case,
        {
            "file_count": 28,
            "file_categories": {
                "meeting_screenshot": 15,
                "coordination_sms": 5,
                "sign_in_record": 0,
                "a1_meeting_export": 1,
                "observation_confirmation": 3,
                "presentation_material": 2,
                "meeting_agenda": 1,
                "speaker_profile": 1,
            },
            "files": [
                {
                    "file_id": "file-sign-1",
                    "file_name": "Remote_A1P260307357_20260506_签到表 (1).jpg",
                    "document_category": "meeting_screenshot",
                    "file_type": "image",
                    "confidence": 0.4,
                    "parse_status": "done",
                }
            ],
            "facts": {
                "actual_sign_in_count": 0,
                "planned_attendees": 6,
                "attendance_delta": 6,
                "material_code": "M-CN-00013658",
                "presentation_topic": "SALWEEN研究：重塑PCV初治患者一线新标准",
            },
            "template": {"columns": 143, "filled_row2": 120, "file_name": "固定模板输出.xlsx"},
            "outputs": ["fixed_template_excel", "finding_excel"],
            "outputs_detail": [
                {"output_type": "fixed_template_excel", "file_name": "固定模板输出.xlsx"},
                {"output_type": "finding_excel", "file_name": "Finding.xlsx"},
            ],
        },
    )

    failed = {item["check_id"]: item for item in report["checks"] if not item["passed"]}

    category_diag = failed["category:sign_in_record"]["diagnosis"]
    assert category_diag["stage"] == "classification"
    assert category_diag["target_category"] == "sign_in_record"
    assert category_diag["candidate_files"][0]["file_name"] == "Remote_A1P260307357_20260506_签到表 (1).jpg"
    assert category_diag["candidate_files"][0]["current_category"] == "meeting_screenshot"

    fact_diag = failed["fact:actual_sign_in_count"]["diagnosis"]
    assert fact_diag["stage"] == "fact_extraction"
    assert fact_diag["field"] == "actual_sign_in_count"
    assert fact_diag["source_path"] == "meeting.state_json.meeting_case.actual_sign_in_count"
    assert fact_diag["related_files"][0]["file_id"] == "file-sign-1"

    template_diag = failed["template:filled_row2"]["diagnosis"]
    assert template_diag["stage"] == "template_generation"
    assert template_diag["output_type"] == "fixed_template_excel"
    assert template_diag["file_name"] == "固定模板输出.xlsx"

    output_diag = failed["output:deliverable_package"]["diagnosis"]
    assert output_diag["stage"] == "delivery_scope"
    assert output_diag["missing_output_type"] == "deliverable_package"

    loose_diag = failed["output_scope:no_loose_outputs"]["diagnosis"]
    assert loose_diag["stage"] == "delivery_scope"
    assert loose_diag["loose_outputs"][0]["output_type"] == "finding_excel"


def test_build_db_compliance_snapshot_reads_case_outputs_and_template(db, tmp_path) -> None:
    project = Project(name="评估快照项目", status="completed")
    db.add(project)
    db.commit()
    meeting = Meeting(
        project_id=project.id,
        meeting_code="A1P260307357",
        status="completed",
        state_json={
            "meeting_case": {
                "meeting_code": "A1P260307357",
                "actual_sign_in_count": 7,
                "planned_attendees": 6,
                "attendance_delta": 1,
                "material_code": "M-CN-00013658",
            }
        },
    )
    db.add(meeting)
    db.commit()

    for idx in range(3):
        db.add(
            FileRecord(
                project_id=project.id,
                meeting_id=meeting.id,
                file_name=f"签到表 {idx}.jpg",
                file_type="image",
                document_category="sign_in_record",
                storage_path=f"/tmp/sign-{idx}.jpg",
                parse_status="done",
            )
        )

    template_path = tmp_path / "fixed.xlsx"
    wb = Workbook()
    ws = wb.active
    for col in range(1, 144):
        ws.cell(1, col, f"字段{col}")
        ws.cell(2, col, f"值{col}")
    wb.save(template_path)

    db.add(
        Output(
            project_id=project.id,
            meeting_id=meeting.id,
            output_type="fixed_template_excel",
            file_name="固定模板输出.xlsx",
            storage_path=str(template_path),
        )
    )
    db.add(
        Output(
            project_id=project.id,
            meeting_id=meeting.id,
            output_type="deliverable_package",
            file_name="case.zip",
            storage_path=str(tmp_path / "case.zip"),
        )
    )
    db.commit()

    snapshot = build_db_compliance_snapshot(db, project.id, meeting.id)

    assert snapshot["file_count"] == 3
    assert snapshot["file_categories"]["sign_in_record"] == 3
    assert snapshot["files"][0]["file_name"] == "签到表 0.jpg"
    assert snapshot["files"][0]["document_category"] == "sign_in_record"
    assert snapshot["facts"]["actual_sign_in_count"] == 7
    assert snapshot["template"] == {"columns": 143, "filled_row2": 143}
    assert snapshot["outputs"] == ["deliverable_package", "fixed_template_excel"]
    assert snapshot["outputs_detail"][0]["output_type"] == "fixed_template_excel"

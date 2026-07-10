from __future__ import annotations

from openpyxl import load_workbook

from app.services.outputs.material_layout_deliverables import generate_material_parse_index_excel


def test_material_parse_index_exposes_vision_review_gate(tmp_path) -> None:
    output = tmp_path / "资料解析索引.xlsx"
    generate_material_parse_index_excel(
        [
            {
                "file_name": "到场确认单.png",
                "document_category": "observation_confirmation",
                "ocr_engine": "vision:glm-ocr",
                "vision_confidence": 0.76,
                "vision_quality_score": 0.44,
                "vision_manual_review_required": True,
                "vision_review_reasons": ["low_resolution", "handwriting_risk"],
                "vision_consensus": {
                    "status": "needs_review",
                    "conflicts": [{"field": "actual_sign_in_count", "values": [{"value": 5}, {"value": 6}]}],
                },
                "field_confidence": {"meeting_code": 0.45, "speaker_service_minutes": 0.45},
                "layout_counts": {},
                "char_count": 18,
                "text_content": "现场确认单 医生手写签名",
            }
        ],
        output,
    )

    wb = load_workbook(output)
    ws = wb["资料解析索引"]
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]
    values = dict(zip(headers, row))

    assert "视觉质量分" in headers
    assert "需识别复核" in headers
    assert "复核原因" in headers
    assert "共识状态" in headers
    assert "共识冲突" in headers
    assert "字段置信度" in headers
    assert values["需识别复核"] == "是"
    assert "handwriting_risk" in values["复核原因"]
    assert values["共识状态"] == "needs_review"
    assert "actual_sign_in_count" in values["共识冲突"]
    assert "meeting_code: 0.45" in values["字段置信度"]

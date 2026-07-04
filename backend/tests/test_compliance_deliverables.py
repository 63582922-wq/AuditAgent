from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from app.services.outputs.compliance_deliverables import (
    build_compliance_deliverable_bundle,
    generate_fixed_template_excel,
)


def test_build_compliance_deliverable_bundle(tmp_path):
    findings = [
        {
            "risk_id": "FIND-001",
            "risk_level": "高",
            "risk_score": 85,
            "risk_category": "计划不一致",
            "problem": "讲者时长超出计划",
            "rule_triggered": "CMP-001",
            "evidence_json": {"speaker_service_minutes": 45},
            "suggestion": "核实 A1 计划并补充说明",
            "manual_review_required": False,
            "status": "pending",
            "source_file_id": "file-1",
        }
    ]
    missing = [
        {"document_type": "sign_in_record", "importance": "中", "reason": "缺少签到记录"},
    ]
    materials = [
        {
            "file_id": "file-1",
            "file_name": "签到表.jpg",
            "document_category": "sign_in_record",
            "ocr_engine": "vision:glm-ocr",
            "text_content": "签到人数 12",
            "md_results": "# 签到表\n\n签到人数 12",
            "layout_details": [[{"label": "text", "content": "签到人数 12"}]],
            "layout_counts": {"text": 1, "table": 0, "image": 0, "formula": 0, "other": 0},
            "char_count": 8,
        }
    ]
    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 TEST",
        findings,
        missing,
        meeting_case={"meeting_code": "A1PTEST001", "observation_type": "远程观察"},
        file_names={"file-1": "签到表.jpg"},
        parsed_materials=materials,
    )

    assert bundle["deliverable_package"].exists()
    assert bundle["fixed_template_excel"].exists()
    assert bundle["finding_pdf"].suffix == ".pdf"
    assert bundle["finding_excel"].suffix == ".xlsx"
    assert bundle["observation_summary"].exists()
    assert bundle["evidence_index"].exists()
    assert bundle["fixed_template_field_evidence"].exists()
    assert bundle["material_parse_index"].exists()
    assert bundle["missing_docs"].exists()
    assert bundle["correction_list"].exists()
    assert not (bundle["material_parse_index"].parent / "markdown").exists()
    assert not (bundle["material_parse_index"].parent / "layout").exists()
    assert "RemoteObservation" in bundle["deliverable_package"].name
    with ZipFile(bundle["deliverable_package"]) as zf:
        package_names = zf.namelist()
    assert not any("/markdown/" in name or "/layout/" in name for name in package_names)

    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[2]]
    values = [c.value for c in ws[3]]
    row = dict(zip(headers, values))
    assert row["会议编码"] == "A1PTEST001"
    assert row["观察点汇总（根据前面所选finding填写描述，需逐条写明问题点标题）"]
    assert "补充签到记录" in row["待跟进事项（无法在跟会现场或观察员收集的evidence不足以判断是否为finding，需进一步核实）"]
    wb.close()


def test_fixed_template_uses_sms_watch_excel_fields(tmp_path):
    materials = [
        {
            "file_id": "watch-1",
            "file_name": "Remote_SMS202606090070_20260615_线上直播观看数据.xlsx",
            "document_category": "sign_in_record",
            "sheets": [
                {
                    "sheet_name": "观看记录详情",
                    "header_row": 2,
                    "pre_header_rows": [["赋能术规范-老友技系列项目 乳腺外科点对点手术交流沙龙 6月份线上直播观看数据"]],
                    "columns": [
                        {"name": "会议时间"},
                        {"name": "会议主席"},
                        {"name": "观众姓名"},
                        {"name": "登录时间"},
                        {"name": "登录时长"},
                    ],
                    "rows": [
                        {
                            "row_number": 3,
                            "values": {
                                "会议时间": "2026-06-15 19:00-21:17",
                                "会议主席": "邓甬川",
                                "观众姓名": "李*",
                                "登录时间": "2026-06-15 18:04:47",
                                "登录时长": "00:52:48",
                            },
                        },
                        {
                            "row_number": 4,
                            "values": {
                                "会议时间": "2026-06-15 19:00-21:17",
                                "会议主席": "邓甬川",
                                "观众姓名": "张*",
                                "登录时间": "2026-06-15 18:12:47",
                                "登录时长": "01:10:51",
                            },
                        },
                    ],
                }
            ],
        }
    ]
    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 SMS",
        [],
        [],
        meeting_case={
            "meeting_code": "SMS202606090070",
            "source_folder": "Remote_SMS202606090070_20260615_Lei, Lily Yuli_Supporting",
        },
        parsed_materials=materials,
    )

    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    row = dict(zip([c.value for c in ws[2]], [c.value for c in ws[3]]))
    assert row["会议编码"] == "SMS202606090070"
    assert row["观察类型"] == "远程观察"
    assert str(row["实际会议日期"])[:10] == "2026-06-15"
    assert str(row["实际会议开始时间"]) in {"19:00:00", "19:00"}
    assert str(row["实际会议结束时间"]) in {"21:17:00", "21:17"}
    assert "乳腺外科点对点手术交流沙龙" in row["主题"]
    assert row["开始时人数（不含Roche员工）"] == "待补充/需核实"
    assert row["会中最大人数（不含Roche员工）"] == "2（观看记录行数，需端口截图核实）"
    assert row["结束时人数（不含Roche员工）"] == "待补充/需核实"
    assert row["签到表类型（电子/纸质/电子+纸质/无签到表）"] == "N/A"
    assert "核实会中最大人数" in row["待跟进事项（无法在跟会现场或观察员收集的evidence不足以判断是否为finding，需进一步核实）"]
    wb.close()

    evidence_wb = load_workbook(bundle["fixed_template_field_evidence"], data_only=True)
    evidence_ws = evidence_wb.active
    evidence_rows = list(evidence_ws.iter_rows(values_only=True))
    assert any(r[1] == "实际会议日期" and "资料表格" in str(r[4]) for r in evidence_rows)
    assert any(r[1] == "会中最大人数（不含Roche员工）" and r[3] == "needs_review" for r in evidence_rows)
    evidence_wb.close()


def test_fixed_template_uses_vision_material_fields(tmp_path):
    materials = [
        {
            "file_id": "ppt-1",
            "file_name": "Remote_SMS202606090070_20260615_PPT.jpg",
            "document_category": "presentation_material",
            "fields": {
                "material_code": "P-HPK-2025.05-090 Valid Until 2027.05",
                "presentation_topic": "新剂型-助力HER2阳性晚期一线走向高质量治愈",
                "ppt_pages": 30,
                "speaker_name": "周美琪",
                "vision_confidence": 0.91,
            },
            "field_confidence": {"material_code": 0.91, "ppt_pages": 0.88},
            "text_content": "PPT主题 新剂型-助力HER2阳性晚期一线走向高质量治愈 P-HPK-2025.05-090 Valid Until 2027.05 共30页",
        },
        {
            "file_id": "shot-1",
            "file_name": "Remote_SMS202606090070_20260615_最大端口数_zoom端 (2).jpg",
            "document_category": "meeting_screenshot",
            "fields": {
                "actual_platform": "ZOOM 95496290261",
                "start_attendee_count": "7+46人次",
                "max_attendee_count": "5+61人次",
                "end_attendee_count": "8+61人次",
                "vision_confidence": 0.84,
            },
            "field_confidence": {"max_attendee_count": 0.84},
            "text_content": "ZOOM 95496290261 最大端口数 5+61人次",
        },
        {
            "file_id": "other-1",
            "file_name": "Remote_SMS202606090070_20260615_其他厂家.jpg",
            "document_category": "other_supporting_evidence",
            "fields": {
                "other_company_seen": "是",
                "other_company_name": "诺华",
                "actual_sponsor": "中国医学基金会",
                "vision_confidence": 0.83,
            },
            "field_confidence": {"other_company_seen": 0.83},
            "text_content": "观察到其他厂家 logo：诺华；主办方：中国医学基金会",
        },
    ]

    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 SMS Vision",
        [],
        [],
        meeting_case={
            "meeting_code": "SMS202606090070",
            "source_folder": "Remote_SMS202606090070_20260615_Lei, Lily Yuli_Supporting",
            "observation_type": "远程观察",
        },
        parsed_materials=materials,
    )

    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    row = dict(zip([c.value for c in ws[2]], [c.value for c in ws[3]]))
    assert row["实际会议地点（线上平台）"] == "ZOOM 95496290261"
    assert row["开始时人数（不含Roche员工）"] == "7+46人次"
    assert row["会中最大人数（不含Roche员工）"] == "5+61人次"
    assert row["结束时人数（不含Roche员工）"] == "8+61人次"
    assert row["PPT页数"] == 30
    assert "P-HPK-2025.05-090" in row["PPT主题及编码"]
    assert row["实际主办方\n(如无的话写N/A）"] == "中国医学基金会"
    assert row["会议现场出现其他厂商员工"] == 1
    assert row["是否问题会议\n（前面问题点选项有1的，此处填1，若无，填0）"] == 1
    wb.close()


def test_fixed_template_prefers_confirmation_facts_over_watch_rows_and_low_confidence_ppt(tmp_path):
    materials = [
        {
            "file_id": "watch-1",
            "file_name": "Remote_SMS202606090070_20260615_线上直播观看数据.xlsx",
            "document_category": "sign_in_record",
            "sheets": [
                {
                    "sheet_name": "观看记录详情",
                    "header_row": 2,
                    "pre_header_rows": [["赋能术规范-老友“技”系列项目 乳腺外科点对点手术交流沙龙 6月份线上直播观看数据"]],
                    "columns": [{"name": "会议时间"}, {"name": "会议主席"}, {"name": "观众姓名"}],
                    "rows": [
                        {
                            "row_number": 3,
                            "values": {
                                "会议时间": "2026-06-15 19:00-21:17",
                                "会议主席": "邓甬川",
                                "观众姓名": "李*",
                            },
                        }
                    ],
                }
            ],
        },
        {
            "file_id": "confirm-1",
            "file_name": "Remote_SMS202606090070_20260615_确认单 (1).jpg",
            "document_category": "observation_confirmation",
            "text_content": (
                "2026年6月15日线上会议（ZOOM），实际开始19:00，结束21:23，时长143分钟。"
                "实际会议开始人数7名参会者+46人次，实际会议最大人数5名参会者+61人次，"
                "实际会议结束人数8名参会者+61人次。"
            ),
            "fields": {
                "summary_text": "实际开始19:00，结束21:23；开始7+46人次，最大5+61人次，结束8+61人次。",
                "vision_confidence": 0.84,
                "meeting_code": "SMS202606090070",
                "observation_success": "是",
            },
            "field_confidence": {"meeting_code": 0.84, "observation_success": 0.84},
        },
        {
            "file_id": "ppt-1",
            "file_name": "Remote_SMS202606090070_20260615_PPT.jpg",
            "document_category": "presentation_material",
            "text_content": "标题为新剂型-动力HER2阳性晚期一线高质量治愈，底部 P-HPK-2025.05-090 Valid Until 2027.05，状态栏 幻灯片 1/30",
            "fields": {
                "presentation_topic": "新剂型动力HER2阳性晚期一线高质量治愈",
                "material_code": "P-HPK-2025.05-090",
                "ppt_pages": 1,
                "vision_confidence": 0.45,
            },
            "field_confidence": {"presentation_topic": 0.45, "ppt_pages": 0.45},
        },
        {
            "file_id": "sponsor-1",
            "file_name": "Remote_SMS202606090070_20260615_赞助回报_专题会.jpg",
            "document_category": "other_supporting_evidence",
            "text_content": "线上Zoom会议截图，会议标题为“新剂型-助力HER2阳性晚期一线走向高质量治愈”，讲者为周美琪。",
            "fields": {
                "summary_text": "会议标题为“新剂型-助力HER2阳性晚期一线走向高质量治愈”。",
                "vision_confidence": 0.9,
            },
            "field_confidence": {},
        },
    ]

    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 SMS Confirmation",
        [],
        [],
        meeting_case={
            "meeting_code": "SMS202606090070",
            "source_folder": "Remote_SMS202606090070_20260615_Lei, Lily Yuli_Supporting",
            "observation_type": "远程观察",
        },
        parsed_materials=materials,
    )

    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    row = dict(zip([c.value for c in ws[2]], [c.value for c in ws[3]]))
    assert str(row["实际会议结束时间"]) in {"21:23:00", "21:23"}
    assert row["开始时人数（不含Roche员工）"] == "7+46人次"
    assert row["会中最大人数（不含Roche员工）"] == "5+61人次"
    assert row["结束时人数（不含Roche员工）"] == "8+61人次"
    assert row["PPT页数"] == 30
    assert "新剂型-助力HER2阳性晚期一线走向高质量治愈" in row["PPT主题及编码"]
    assert "动力" not in row["PPT主题及编码"]
    wb.close()


def test_fixed_template_calibration_row_overrides_low_confidence_fields_but_not_findings(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = load_workbook(Path("FX/新建 Microsoft Excel 工作表.xlsx"))
    ws = wb.active
    headers = [c.value for c in ws[2]]
    data = [None for _ in headers]
    for idx, header in enumerate(headers):
        if header == "会议编码":
            data[idx] = "SMS-CAL-001"
        elif header == "PPT页数":
            data[idx] = 30
        elif header == "PPT主题及编码":
            data[idx] = "新剂型-助力HER2阳性晚期一线走向高质量治愈\nP-HPK-2025.05-090 Valid Until 2027.05"
        elif header == "是否问题会议\n（前面问题点选项有1的，此处填1，若无，填0）":
            data[idx] = 0
    ws.append(data)
    wb.save(template)
    wb.close()

    output = tmp_path / "out.xlsx"
    results = generate_fixed_template_excel(
        {
            "meeting_code": "SMS-CAL-001",
            "source_folder": "/tmp/FX/Remote_SMS-CAL-001_Supporting",
            "observation_type": "远程观察",
        },
        findings=[
            {
                "problem": "会议实际会议日程与计划不符",
                "risk_category": "计划不一致",
                "manual_review_required": False,
            }
        ],
        missing=[],
        output=output,
        parsed_materials=[
            {
                "file_name": "PPT.jpg",
                "document_category": "presentation_material",
                "fields": {
                    "presentation_topic": "新剂型动力HER2阳性晚期一线高质量治愈",
                    "material_code": "P-HPK-2025.05-090",
                    "ppt_pages": 1,
                    "vision_confidence": 0.45,
                },
            }
        ],
        template_path=template,
    )

    result_by_header = {r.header: r for r in results}
    assert result_by_header["PPT页数"].status == "calibrated"
    assert result_by_header["PPT页数"].value == 30
    assert "助力" in result_by_header["PPT主题及编码"].value
    assert result_by_header["是否问题会议\n（前面问题点选项有1的，此处填1，若无，填0）"].value == 1


def test_fixed_template_rejects_invalid_case_material_code_when_valid_material_exists(tmp_path):
    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 SMS material code",
        [],
        [],
        meeting_case={
            "meeting_code": "SMS202606090070",
            "observation_type": "远程观察",
            "material_code": "Promotional。",
        },
        parsed_materials=[
            {
                "file_id": "ppt-1",
                "file_name": "Remote_SMS202606090070_20260615_PPT.jpg",
                "document_category": "presentation_material",
                "fields": {
                    "material_code": "P-HPK-2025.05-090",
                    "presentation_topic": "新剂型-助力HER2阳性晚期一线走向高质量治愈",
                    "vision_confidence": 0.95,
                },
                "field_confidence": {"material_code": 0.95},
                "text_content": "页脚编码 P-HPK-2025.05-090 Valid Until 2027.05",
            }
        ],
    )

    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    row = dict(zip([c.value for c in ws[2]], [c.value for c in ws[3]]))
    assert "P-HPK-2025.05-090" in row["PPT主题及编码"]
    assert "Promotional。" not in row["PPT主题及编码"]
    wb.close()

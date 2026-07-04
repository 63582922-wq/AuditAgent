from __future__ import annotations

import json
from zipfile import ZipFile

from openpyxl import load_workbook

from app.models import Meeting, Project
from app.services.agent.workflow import AgentWorkflow
from app.services.domain.compliance.template_field_engine import TemplateFieldResult
from app.services.outputs.compliance_deliverables import build_compliance_deliverable_bundle
from app.services.outputs.template_quality import evaluate_template_field_results


def test_template_quality_summarizes_field_statuses() -> None:
    report = evaluate_template_field_results(
        [
            TemplateFieldResult(1, "会议编码", "SMS202606090070", "extracted", "资料", 0.92, "source"),
            TemplateFieldResult(2, "会中最大人数", "2（观看记录行数，需端口截图核实）", "needs_review", "资料表格", 0.42, "watch rows"),
            TemplateFieldResult(3, "计划预算金额", "待补充/需核实", "missing", "字段审计引擎", 0.0, "missing"),
            TemplateFieldResult(4, "内部订单号", "待PMO填写", "manual_required", "PMO字段", 0.0, "pmo"),
            TemplateFieldResult(5, "Roche comments", "待客户确认", "customer_required", "客户字段", 0.0, "customer"),
            TemplateFieldResult(6, "", None, "blank_header", "模板", 0.0, ""),
        ],
        expected_field_count=6,
    )

    assert report["total_fields"] == 6
    assert report["assessed_fields"] == 6
    assert report["status"] == "fail"
    assert report["counts"]["complete"] == 1
    assert report["counts"]["needs_review"] == 1
    assert report["counts"]["missing"] == 1
    assert report["counts"]["manual_required"] == 1
    assert report["counts"]["customer_required"] == 1
    assert report["counts"]["structural_error"] == 1
    assert report["owner_counts"]["system"] == 2
    assert report["owner_counts"]["observer"] == 2
    assert report["owner_counts"]["pmo"] == 1
    assert report["owner_counts"]["customer"] == 1
    assert report["review_required"] is True
    assert report["items"][0]["owner"] == "system"
    assert report["items"][0]["evidence_type"] == "direct_material"
    assert report["items"][1]["quality"] == "needs_review"
    assert report["items"][1]["owner"] == "observer"
    assert report["items"][1]["handoff_required"] is True
    assert report["items"][3]["owner"] == "pmo"
    assert report["items"][4]["owner"] == "customer"
    assert report["items"][5]["quality"] == "structural_error"


def test_deliverable_bundle_includes_template_quality_gate(tmp_path) -> None:
    bundle = build_compliance_deliverable_bundle(
        tmp_path,
        "观察案件 SMS",
        [],
        [],
        meeting_case={"meeting_code": "SMS202606090070", "observation_type": "远程观察"},
        parsed_materials=[
            {
                "file_id": "shot-1",
                "file_name": "Remote_SMS202606090070_20260615_最大端口数_zoom端.jpg",
                "document_category": "meeting_screenshot",
                "fields": {
                    "actual_platform": "ZOOM 95496290261",
                    "start_attendee_count": "7+46人次",
                    "max_attendee_count": "5+61人次",
                    "end_attendee_count": "8+61人次",
                    "vision_confidence": 0.84,
                },
                "text_content": "ZOOM 95496290261 最大端口数 5+61人次",
            }
        ],
    )

    quality_xlsx = bundle["fixed_template_quality"]
    quality_json = bundle["fixed_template_quality_json"]
    assert quality_xlsx.exists()
    assert quality_json.exists()

    summary = json.loads(quality_json.read_text(encoding="utf-8"))
    wb = load_workbook(bundle["fixed_template_excel"], data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[2] if cell.value not in (None, "")]
    wb.close()

    assert len(headers) == 143
    assert summary["total_fields"] == len(headers)
    assert summary["assessed_fields"] == len(headers)
    assert summary["status"] in {"pass", "needs_review"}
    assert summary["counts"]["complete"] > 0
    assert summary["counts"]["missing"] > 0

    report_wb = load_workbook(quality_xlsx, data_only=True)
    report_ws = report_wb.active
    assert report_ws["A1"].value == "固定模板质量门禁"
    assert any(row[1] == "会议编码" and row[4] == "complete" for row in report_ws.iter_rows(values_only=True))
    report_wb.close()

    with ZipFile(bundle["deliverable_package"]) as zf:
        names = zf.namelist()
    assert any(name.endswith("00_固定模板/固定模板质量门禁.xlsx") for name in names)
    assert any(name.endswith("00_固定模板/固定模板质量门禁.json") for name in names)


def test_workflow_writes_template_quality_into_meeting_deliverable_json(db, tmp_path, monkeypatch) -> None:
    from app.config import settings

    monkeypatch.setattr(settings, "storage_path", tmp_path)

    project = Project(
        name="SMS 质量门禁项目",
        status="completed",
        state_json={"agent_domain": "compliance"},
    )
    db.add(project)
    db.commit()
    meeting = Meeting(
        project_id=project.id,
        meeting_code="SMS202606090070",
        status="completed",
        state_json={
            "agent_domain": "compliance",
            "meeting_case": {"meeting_code": "SMS202606090070", "observation_type": "远程观察"},
        },
        deliverable_json={"status": "pending", "comment": ""},
    )
    db.add(meeting)
    db.commit()

    workflow = AgentWorkflow(db, project.id, meeting_id=meeting.id)
    workflow._generate_outputs(project, [])
    db.refresh(meeting)

    quality = (meeting.deliverable_json or {}).get("template_quality")
    assert quality
    assert quality["total_fields"] == 143
    assert quality["assessed_fields"] == 143
    assert quality["status"] in {"pass", "needs_review"}
    assert quality["counts"]["missing"] > 0

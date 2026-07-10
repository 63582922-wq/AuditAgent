from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook
import pytest

from app.models import AgentRunLog, FileRecord, Meeting, Output, Project
from app.services.domain.compliance.case_loader import import_case_folder
from app.services.domain.compliance.cross_checker import run_compliance_checks
from app.services.cross_checker import check_missing_documents
from app.services.domain.compliance.classifier import classify_compliance_document, infer_meeting_code


FX_ROOT = Path(__file__).resolve().parents[2] / "FX"
FX_CASE = FX_ROOT / "Remote_A1P260307357_20260506_Luo, Amy Yun_Supporting"


def test_classify_a1_export():
    cls = classify_compliance_document("Remote_A1P260307357_export.pdf", ".pdf", "会议编号 A1P260307357")
    assert cls["document_category"] == "a1_meeting_export"
    assert cls["confidence"] >= 0.9


def test_infer_meeting_code():
    code = infer_meeting_code(["Remote_A1P260307357_20260506.pdf", "other.xlsx"])
    assert code == "A1P260307357"
    sms_code = infer_meeting_code(["Remote_SMS202606090070_20260615_确认单.jpg"])
    assert sms_code == "SMS202606090070"


def test_import_case_folder(db, monkeypatch):
    if not FX_CASE.is_dir():
        pytest.skip("FX 样本目录不存在")

    monkeypatch.setattr("app.config.settings.agent_domain", "compliance")
    project_id, meeting_id, profile = import_case_folder(db, FX_CASE, "测试案件")
    assert profile.get("meeting_code") == "A1P260307357"
    assert project_id
    assert meeting_id
    assert profile.get("source_folder", "").endswith("Remote_A1P260307357_20260506_Luo, Amy Yun_Supporting")


def test_import_case_folder_does_not_use_full_pdf_parser(db, monkeypatch):
    if not FX_CASE.is_dir():
        pytest.skip("FX 样本目录不存在")

    calls = []

    def record_parse_pdf(path):
        calls.append(Path(path).name)
        return {"text_content": ""}

    monkeypatch.setattr("app.config.settings.agent_domain", "compliance")
    monkeypatch.setattr("app.services.parsers.pdf_parser.parse_pdf", record_parse_pdf)

    project_id, meeting_id, profile = import_case_folder(db, FX_CASE, "轻量导入测试")

    assert project_id
    assert meeting_id
    assert profile.get("meeting_code") == "A1P260307357"
    assert calls == []


def test_import_case_folder_rejects_multi_case_fx_root(db):
    if not FX_ROOT.is_dir():
        pytest.skip("FX 样本目录不存在")

    with pytest.raises(ValueError, match="单场观察案件"):
        import_case_folder(db, FX_ROOT, "错误导入")


def test_compliance_rules_on_facts():
    facts = {
        "meeting_code": "A1P260307357",
        "speaker_service_minutes": 10,
        "planned_duration_minutes": 30,
        "actual_duration_minutes": 30,
        "duration_delta_minutes": 0,
        "has_confirmation": True,
        "has_a1_export": True,
        "has_coordination_sms": True,
        "material_code": "M-CN-00013658",
        "planned_attendees": 7,
        "actual_sign_in_count": 7,
        "attendance_delta": 0,
    }
    rules = [
        {
            "rule_id": "CMP-001",
            "rule_name": "付费讲者讲课时长不足15分钟",
            "risk_category": "违反公司制度",
            "risk_level": "高",
            "condition": {"all": [{"field": "speaker_service_minutes", "operator": "<", "value": 15}]},
            "evidence_fields": ["speaker_service_minutes"],
            "suggestion_template": "时长不足",
            "enabled": True,
        }
    ]
    hits = run_compliance_checks(facts, rules)
    assert len(hits) == 1
    assert hits[0]["rule_triggered"] == "CMP-001"


def test_sms_rules_do_not_require_a1_or_treat_unparsed_material_as_finding():
    missing = check_missing_documents(
        {"sign_in_record", "observation_confirmation", "presentation_material", "coordination_sms"},
        domain="compliance",
        meeting_case={
            "meeting_code": "SMS202606090070",
            "observation_type": "远程观察",
        },
    )
    assert all(item["document_type"] != "a1_meeting_export" for item in missing)
    assert all(item["document_type"] != "meeting_metadata" for item in missing)

    facts = {
        "meeting_code": "SMS202606090070",
        "has_a1_export": False,
        "has_presentation_material": True,
        "material_code": "",
        "material_code_pending_vision": True,
        "attendance_source": "watch_record",
        "planned_attendees": 50,
        "actual_sign_in_count": 0,
        "attendance_delta": 50,
    }
    rules = [
        {
            "rule_id": "CMP-004",
            "rule_name": "缺少A1会议导出",
            "risk_category": "缺件",
            "risk_level": "高",
            "condition": {"all": [{"field": "has_a1_export", "operator": "==", "value": "false"}]},
            "suggestion_template": "补充 A1",
            "enabled": True,
        },
        {
            "rule_id": "CMP-005",
            "rule_name": "演讲材料编码缺失",
            "risk_category": "材料",
            "risk_level": "中",
            "condition": {"all": [{"field": "material_code", "operator": "is_empty"}]},
            "suggestion_template": "核实材料编码",
            "enabled": True,
        },
        {
            "rule_id": "CMP-006",
            "rule_name": "签到人数与计划不符",
            "risk_category": "参会",
            "risk_level": "中",
            "condition": {"all": [{"field": "attendance_delta", "operator": ">", "value": 2}]},
            "suggestion_template": "核实人数",
            "enabled": True,
        },
    ]
    assert run_compliance_checks(facts, rules) == []


def test_harness_import_and_rules_only(db, monkeypatch):
    if not FX_CASE.is_dir():
        pytest.skip("FX 样本目录不存在")

    from app.services.agent.harness import ComplianceHarness

    monkeypatch.setattr("app.config.settings.agent_domain", "compliance")
    monkeypatch.setattr(
        "app.services.domain.compliance.finding_generator.generate_finding_narratives",
        lambda hits, profile, obs_type="Remote": [{**h, "analysis": h["suggestion"]} for h in hits],
    )

    harness = ComplianceHarness(db)
    project_id, meeting_id, _ = harness.import_case(FX_CASE, "Harness 测试")
    result = harness.run(project_id, meeting_id, skip_orchestrator=True)
    assert result.meeting_code == "A1P260307357"
    assert result.finding_count >= 0
    assert result.status in ("completed", "needs_review", "accepted")


def test_harness_passes_meeting_id_to_orchestrator(db, monkeypatch, tmp_path):
    from app.services.agent.harness import compliance_harness
    from app.services.agent.harness import ComplianceHarness

    monkeypatch.setattr("app.config.settings.agent_domain", "compliance")

    project = Project(name="Harness meeting scope", status="created")
    db.add(project)
    db.commit()
    meeting = Meeting(
        project_id=project.id,
        meeting_code="SMS202606090070",
        observation_type="远程观察",
        status="draft",
        state_json={"meeting_case": {"meeting_code": "SMS202606090070", "observation_type": "远程观察"}},
    )
    db.add(meeting)
    db.commit()
    file_path = tmp_path / "Remote_SMS202606090070_20260615_线上直播观看数据.xlsx"
    file_path.write_bytes(b"placeholder")
    db.add(
        FileRecord(
            project_id=project.id,
            meeting_id=meeting.id,
            file_name=file_path.name,
            file_type="excel",
            document_category="sign_in_record",
            storage_path=str(file_path),
            parse_status="uploaded",
        )
    )
    db.commit()

    captured = {}

    class FakeMissionOrchestrator:
        def __init__(self, db_arg, project_id, progress_callback=None, trace=None, meeting_id=None):
            captured["meeting_id"] = meeting_id

        def run(self):
            return {"mode": "orchestrator"}

    class FakeRuntime:
        def __init__(self, *args, **kwargs):
            self.trace = None

        def _post_process(self, mode):
            return SimpleNamespace(status="completed", critic_summary=None, human_gate=None)

    monkeypatch.setattr(compliance_harness, "MissionOrchestrator", FakeMissionOrchestrator)
    monkeypatch.setattr(compliance_harness, "AgentRuntime", FakeRuntime)
    monkeypatch.setattr(ComplianceHarness, "_reclassify_files", lambda self, project_id, meeting_id: None)
    monkeypatch.setattr(ComplianceHarness, "_ensure_parsed_documents", lambda self, project_id, meeting_id: 0)
    monkeypatch.setattr(ComplianceHarness, "_run_compliance_rules", lambda self, project_id, meeting_id: [])
    monkeypatch.setattr(ComplianceHarness, "_persist_findings", lambda self, project_id, meeting_id, findings, files=None: None)
    monkeypatch.setattr(ComplianceHarness, "_sync_meeting_snapshot", lambda self, project_id, meeting_id, findings, missing, files=None: None)

    harness = ComplianceHarness(db)
    harness.run(project.id, meeting.id, skip_orchestrator=False)

    assert captured["meeting_id"] == meeting.id


def test_harness_persists_automatic_evaluation_after_outputs(db, monkeypatch, tmp_path):
    from app.services.agent.harness import compliance_harness
    from app.services.agent.harness import ComplianceHarness

    monkeypatch.setattr("app.config.settings.agent_domain", "compliance")

    project = Project(name="Harness 自动评估", status="created")
    db.add(project)
    db.commit()
    meeting = Meeting(
        project_id=project.id,
        meeting_code="A1P260307357",
        observation_type="远程观察",
        status="draft",
        state_json={
            "meeting_case": {
                "meeting_code": "A1P260307357",
                "actual_sign_in_count": 7,
                "planned_attendees": 6,
                "attendance_delta": 1,
                "material_code": "M-CN-00013658",
                "presentation_topic": "SALWEEN研究：重塑PCV初治患者一线新标准",
            }
        },
    )
    db.add(meeting)
    db.commit()

    category_counts = {
        "meeting_screenshot": 12,
        "coordination_sms": 5,
        "sign_in_record": 3,
        "a1_meeting_export": 1,
        "observation_confirmation": 3,
        "presentation_material": 2,
        "meeting_agenda": 1,
        "speaker_profile": 1,
    }
    for category, count in category_counts.items():
        for idx in range(count):
            db.add(
                FileRecord(
                    project_id=project.id,
                    meeting_id=meeting.id,
                    file_name=f"{category}-{idx}.jpg",
                    file_type="image",
                    document_category=category,
                    storage_path=str(tmp_path / f"{category}-{idx}.jpg"),
                    parse_status="done",
                )
            )
    db.commit()

    template_path = tmp_path / "固定模板输出.xlsx"
    wb = Workbook()
    ws = wb.active
    for col in range(1, 144):
        ws.cell(1, col, f"字段{col}")
        ws.cell(2, col, f"值{col}")
    wb.save(template_path)
    zip_path = tmp_path / "交付包.zip"
    zip_path.write_bytes(b"zip")

    class FakeRuntime:
        def __init__(self, db_arg, project_id_arg, progress_callback=None, meeting_id=None):
            self.db = db_arg
            self.project_id = project_id_arg
            self.meeting_id = meeting_id
            self.trace = None

        def _post_process(self, mode):
            self.db.add(
                Output(
                    project_id=self.project_id,
                    meeting_id=self.meeting_id,
                    output_type="fixed_template_excel",
                    file_name="固定模板输出.xlsx",
                    storage_path=str(template_path),
                )
            )
            self.db.add(
                Output(
                    project_id=self.project_id,
                    meeting_id=self.meeting_id,
                    output_type="deliverable_package",
                    file_name="交付包.zip",
                    storage_path=str(zip_path),
                )
            )
            current = self.db.get(Meeting, self.meeting_id)
            deliverable = {
                "status": "pending",
                "comment": "",
                "template_quality": {"status": "pass", "total_fields": 143},
            }
            state = dict(current.state_json or {})
            state["deliverable"] = deliverable
            current.state_json = state
            current.deliverable_json = deliverable
            self.db.commit()
            return SimpleNamespace(status="completed", critic_summary=None, human_gate=None)

    monkeypatch.setattr(compliance_harness, "AgentRuntime", FakeRuntime)
    monkeypatch.setattr(ComplianceHarness, "_reclassify_files", lambda self, project_id, meeting_id: None)
    monkeypatch.setattr(ComplianceHarness, "_ensure_parsed_documents", lambda self, project_id, meeting_id: 0)
    monkeypatch.setattr(ComplianceHarness, "_run_compliance_rules", lambda self, project_id, meeting_id: [])
    monkeypatch.setattr(ComplianceHarness, "_persist_findings", lambda self, project_id, meeting_id, findings, files=None: None)

    harness = ComplianceHarness(db)
    result = harness.run(project.id, meeting.id, skip_orchestrator=True)

    refreshed = db.get(Meeting, meeting.id)
    assert result.status == "completed"
    assert refreshed.deliverable_json["template_quality"]["status"] == "pass"
    assert refreshed.deliverable_json["evaluation"]["case_id"] == "A1P260307357"
    assert refreshed.deliverable_json["evaluation"]["passed"] is True
    assert refreshed.state_json["evaluation"]["critical_failures"] == 0
    assert (
        db.query(AgentRunLog)
        .filter_by(project_id=project.id, meeting_id=meeting.id, step="compliance_evaluation")
        .count()
        == 1
    )


def test_harness_evaluation_failure_gates_deliverable(db, monkeypatch):
    from app.services.agent.harness import compliance_harness
    from app.services.agent.harness import ComplianceHarness

    project = Project(name="评估失败门禁", status="completed")
    db.add(project)
    db.commit()
    meeting = Meeting(
        project_id=project.id,
        meeting_code="SMS202606090070",
        observation_type="远程观察",
        status="completed",
        state_json={"meeting_case": {"meeting_code": "SMS202606090070"}},
        deliverable_json={"status": "pending", "comment": ""},
    )
    db.add(meeting)
    db.commit()

    def fake_eval(db_arg, project_id_arg, meeting_id_arg):
        return {
            "status": "completed",
            "case_id": "SMS202606090070",
            "case_name": "SMS sample",
            "meeting_code": "SMS202606090070",
            "passed": False,
            "critical_failures": 2,
            "warning_failures": 0,
            "total_checks": 3,
            "passed_checks": 1,
            "checks": [
                {
                    "check_id": "fact:presentation_topic",
                    "passed": False,
                    "severity": "critical",
                    "expected": "新剂型-助力HER2阳性晚期一线走向高质量治愈",
                    "actual": "PIK3CA突变、HR+/HER2-晚期乳腺癌一线病例分享",
                    "message": "presentation_topic 必须来自资料事实链",
                },
                {
                    "check_id": "category:observation_confirmation",
                    "passed": False,
                    "severity": "critical",
                    "expected": 3,
                    "actual": 2,
                    "message": "observation_confirmation 分类数量必须稳定",
                },
            ],
        }

    monkeypatch.setattr(compliance_harness, "run_db_compliance_evaluation", fake_eval)

    harness = ComplianceHarness(db)
    harness.trace = compliance_harness.AgentTrace(db, project.id, meeting.id)
    report = harness._persist_evaluation(project.id, meeting.id)

    refreshed = db.get(Meeting, meeting.id)
    assert report["passed"] is False
    assert refreshed.deliverable_json["status"] == "needs_review"
    assert refreshed.deliverable_json["evaluation_gate"]["reason"] == "automatic_evaluation_failed"
    assert refreshed.deliverable_json["evaluation_gate"]["failed_check_ids"] == [
        "fact:presentation_topic",
        "category:observation_confirmation",
    ]

    harness._emit_progress(project.id, meeting.id, "completed", 100)
    refreshed = db.get(Meeting, meeting.id)
    assert refreshed.status == "needs_review"
    assert refreshed.state_json["runtime_live"]["step"] == "completed"
